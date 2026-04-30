"""
Data Service Layer - API integration with caching and CSV fallback.
Handles: ACLED, EIA (Brent + SPR), yfinance (DXY, OVX), FRED (China BCI)
"""
import bisect
import json
import math
import os
import re
import time
import logging
import threading
from pathlib import Path
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any

import pandas as pd
import numpy as np
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

import config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# yfinance logs every Yahoo rejection at ERROR level, and on Render (or any
# cloud IP without a cached Yahoo consent cookie + crumb) every single call
# fails — `Expecting value: line 1 column 1` is json.loads on an empty/HTML
# body. Those errors are harmless because DXY and OVX have FRED fallbacks and
# Brent has EIA as primary, but they spam the deploy log with scary-looking
# stack traces. Silence yfinance's own logger — we still catch exceptions and
# log a concise WARNING from our own wrapper.
logging.getLogger("yfinance").setLevel(logging.CRITICAL)

# Ensure cache directory exists
config.CACHE_DIR.mkdir(exist_ok=True)

# Browser-like headers to avoid WAF blocks from cloud IPs
_BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
}


# ─── Conflict Theater Location Lookup ────────────────────────────────────────
# Maps lowercase keywords → (lat, lon, canonical_name) for geocoding headlines

CONFLICT_THEATER_LOCATIONS = {
    # Iran
    "tehran": (35.6892, 51.3890, "Tehran, Iran"),
    "isfahan": (32.6546, 51.6680, "Isfahan, Iran"),
    "natanz": (33.5103, 51.9250, "Natanz, Iran"),
    "fordow": (34.7564, 51.0596, "Fordow, Iran"),
    "bushehr": (28.9234, 50.8203, "Bushehr, Iran"),
    "bandar abbas": (27.1865, 56.2808, "Bandar Abbas, Iran"),
    "kharg island": (29.2333, 50.3167, "Kharg Island, Iran"),
    "tabriz": (38.0800, 46.2919, "Tabriz, Iran"),
    "shiraz": (29.5918, 52.5837, "Shiraz, Iran"),
    "qom": (34.6401, 50.8764, "Qom, Iran"),
    "minab": (27.1050, 57.0786, "Minab, Iran"),
    "chabahar": (25.2919, 60.6430, "Chabahar, Iran"),
    "shahran": (35.75, 51.30, "Shahran, Tehran, Iran"),
    "abadan": (30.3392, 48.3043, "Abadan, Iran"),
    # Strait / Gulf
    "strait of hormuz": (26.5667, 56.2500, "Strait of Hormuz"),
    "hormuz": (26.5667, 56.2500, "Strait of Hormuz"),
    "persian gulf": (25.2854, 55.3500, "Persian Gulf"),
    # Oman
    "salalah": (17.0151, 54.0924, "Salalah, Oman"),
    "muscat": (23.5880, 58.3829, "Muscat, Oman"),
    # Gulf States
    "riyadh": (24.7136, 46.6753, "Riyadh, Saudi Arabia"),
    "jeddah": (21.4858, 39.1925, "Jeddah, Saudi Arabia"),
    "shaybah": (22.5167, 54.0000, "Shaybah, Saudi Arabia"),
    "dubai": (25.2048, 55.2708, "Dubai, UAE"),
    "abu dhabi": (24.4539, 54.3773, "Abu Dhabi, UAE"),
    "manama": (26.2285, 50.5860, "Manama, Bahrain"),
    "bahrain": (26.0667, 50.5577, "Bahrain"),
    "doha": (25.2854, 51.5310, "Doha, Qatar"),
    "qatar": (25.2854, 51.5310, "Qatar"),
    "kuwait": (29.3759, 47.9774, "Kuwait City, Kuwait"),
    # Iraq / Levant
    "baghdad": (33.3152, 44.3661, "Baghdad, Iraq"),
    "basra": (30.5085, 47.7804, "Basra, Iraq"),
    "beirut": (33.8938, 35.5018, "Beirut, Lebanon"),
    "damascus": (33.5138, 36.2765, "Damascus, Syria"),
    # Israel
    "tel aviv": (32.0853, 34.7818, "Tel Aviv, Israel"),
    "jerusalem": (31.7683, 35.2137, "Jerusalem, Israel"),
    "haifa": (32.7940, 34.9896, "Haifa, Israel"),
    "ovda": (29.9402, 34.9358, "Ovda Airbase, Israel"),
    # US / West
    "washington": (38.9072, -77.0369, "Washington, DC"),
    "pentagon": (38.8719, -77.0563, "Pentagon, VA"),
    "geneva": (46.2044, 6.1432, "Geneva, Switzerland"),
    # Waterways
    "red sea": (20.0, 38.0, "Red Sea"),
    "suez": (29.9668, 32.5498, "Suez Canal, Egypt"),
    "bab el-mandeb": (12.5833, 43.3333, "Bab el-Mandeb"),
    "bab el mandeb": (12.5833, 43.3333, "Bab el-Mandeb"),
    "gulf of aden": (12.5, 45.0, "Gulf of Aden"),
    # Yemen
    "aden": (12.7855, 45.0187, "Aden, Yemen"),
    "sanaa": (15.3694, 44.1910, "Sanaa, Yemen"),
    "hodeidah": (14.7979, 42.9541, "Hodeidah, Yemen"),
}

# Pre-sorted keys longest-first for greedy matching
_LOCATION_KEYS_SORTED = sorted(CONFLICT_THEATER_LOCATIONS.keys(), key=len, reverse=True)

# ─── Singleflight (concurrent-call coalescing) ───────────────────────────────
#
# When N concurrent requests trigger the same expensive fetch (ACLED bulk,
# Iran events, GDELT, HDX), without a singleflight all N race past the
# memo/cache check and start their own fetches. That produces:
#   - GDELT 429 Too Many Requests
#   - ACLED ReadTimeouts (their API throttles parallel queries)
#   - the "served from JSON fallback while live fetch was still in
#     progress" race in /api/iran-events
#   - 5x peak memory (each parallel fetch holds its own ~70MB blob)
#
# Singleflight ensures only ONE worker actually runs the fetch for a given
# key at a time; concurrent callers attach to the same in-flight result and
# get the same value back when it lands. After completion the entry is
# removed so the next call (post-cache-expiry) triggers a fresh fetch.
#
# Pattern is "leader/waiter": the first thread to register a key becomes the
# leader and runs fn(); subsequent threads wait on the leader's Event and
# receive the leader's result (or its exception).

class _Singleflight:
    """Coalesce concurrent calls keyed by a string. Thread-safe."""

    def __init__(self):
        self._lock = threading.Lock()
        self._calls: Dict[str, Dict[str, Any]] = {}

    def run(self, key: str, fn, timeout: float = 300.0):
        """Run fn() if no other thread is running it for `key`.

        Concurrent callers wait for the leader's result (up to `timeout`
        seconds) and receive the same value. If the leader raises, all
        waiters re-raise the same exception.
        """
        with self._lock:
            cell = self._calls.get(key)
            is_leader = cell is None
            if is_leader:
                cell = {"event": threading.Event(), "value": None, "error": None}
                self._calls[key] = cell

        if is_leader:
            try:
                cell["value"] = fn()
            except BaseException as e:
                cell["error"] = e
            finally:
                # Pop BEFORE waking waiters so any waiter that races to
                # acquire the lock for a *new* call gets a fresh cell.
                with self._lock:
                    self._calls.pop(key, None)
                cell["event"].set()
            if cell["error"] is not None:
                raise cell["error"]
            return cell["value"]
        else:
            ok = cell["event"].wait(timeout=timeout)
            if not ok:
                raise TimeoutError(f"singleflight wait timed out for key={key!r} after {timeout}s")
            if cell["error"] is not None:
                raise cell["error"]
            return cell["value"]


_sf = _Singleflight()


# ─── Cache Helpers ───────────────────────────────────────────────────────────

def _cache_path(key: str) -> Path:
    return config.CACHE_DIR / f"{key}.json"


def _read_cache(key: str, ttl: int) -> Optional[dict]:
    path = _cache_path(key)
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text())
        if time.time() - data.get("_ts", 0) < ttl:
            return data.get("payload")
    except Exception:
        pass
    return None


def _write_cache(key: str, payload):
    path = _cache_path(key)
    data = json.dumps({"_ts": time.time(), "payload": payload}, default=str)
    # Atomic write: write to temp file then rename to avoid corruption on crash
    tmp = str(path) + ".tmp"
    try:
        Path(tmp).write_text(data, encoding="utf-8")
        os.replace(tmp, str(path))
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


# ─── ACLED API ───────────────────────────────────────────────────────────────

_acled_token = None
_acled_token_expires = 0
_acled_token_lock = threading.Lock()
_iran_fetch_error = None  # Store last error for debugging
_acled_fetch_error = None  # Store last ACLED (main) fetch error for debugging
_acled_fetch_source = None  # 'api' | 'cache' | 'fallback' | None
_acled_fetch_ts = 0.0


def _get_acled_token() -> str:
    global _acled_token, _acled_token_expires
    with _acled_token_lock:
        if _acled_token and time.time() < _acled_token_expires:
            return _acled_token

        if not config.ACLED_USERNAME or not config.ACLED_PASSWORD:
            raise ValueError("ACLED credentials not configured (set ACLED_USERNAME and ACLED_PASSWORD env vars)")

        # Token endpoint: bumped from 8s → 20s with one retry. Render's egress
        # to acleddata.com routinely hits 10-15s on cold starts, and an 8s
        # timeout was deterministically falling through to the frozen JSON
        # fallback even when the API was healthy. We can afford to wait for the
        # OAuth handshake — it's a one-time hit per process.
        last_err = None
        resp = None
        for attempt in range(2):
            try:
                resp = requests.post(
                    config.ACLED_TOKEN_URL,
                    data={
                        "username": config.ACLED_USERNAME,
                        "password": config.ACLED_PASSWORD,
                        "grant_type": "password",
                        "client_id": "acled",
                    },
                    headers={**_BROWSER_HEADERS, "Content-Type": "application/x-www-form-urlencoded"},
                    timeout=20,
                )
                last_err = None
                break
            except (requests.Timeout, requests.ConnectionError) as e:
                last_err = e
                if attempt == 0:
                    logger.warning(f"ACLED token: {type(e).__name__}, retrying once")
                    continue
        if last_err is not None or resp is None:
            raise (last_err or RuntimeError("ACLED token request returned no response"))
        if not resp.ok:
            raise requests.HTTPError(
                f"{resp.status_code} for {resp.url} | resp={resp.text[:300]}",
                response=resp,
            )
        try:
            token_data = resp.json()
        except Exception as e:
            raise requests.HTTPError(f"ACLED token response not JSON: {resp.text[:300]}") from e

        if "access_token" not in token_data:
            # Log the response so we can see what ACLED returned
            err_msg = token_data.get("error_description") or token_data.get("error") or str(token_data)[:300]
            raise requests.HTTPError(f"ACLED token response missing access_token: {err_msg}")

        _acled_token = token_data["access_token"]
        _acled_token_expires = time.time() + token_data.get("expires_in", 86400) - 300
        logger.info("ACLED OAuth token acquired")
        return _acled_token


_EVENT_COLUMNS = ["event_id_cnty", "event_date", "event_type", "sub_event_type",
                   "actor1", "actor2", "country", "location", "latitude", "longitude",
                   "notes", "fatalities", "tags", "source", "source_scale"]

_ACLED_FIELDS = "|".join(_EVENT_COLUMNS)
_ACLED_DATE_RANGE = "2023-10-01|2026-12-31"


def _df_to_event_records(df: pd.DataFrame) -> List[dict]:
    """Convert a DataFrame to a list of event dicts (vectorized, no iterrows)."""
    col_map = {c: c.lower() for c in df.columns}
    df = df.rename(columns=col_map).copy()
    str_cols = [c for c in _EVENT_COLUMNS if c not in ("latitude", "longitude", "fatalities")]
    for c in str_cols:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)
    if "latitude" in df.columns:
        df["latitude"] = pd.to_numeric(df["latitude"], errors="coerce")
    if "longitude" in df.columns:
        df["longitude"] = pd.to_numeric(df["longitude"], errors="coerce")
    if "fatalities" in df.columns:
        df["fatalities"] = pd.to_numeric(df["fatalities"], errors="coerce").fillna(0).astype(int)
    records = df[[c for c in _EVENT_COLUMNS if c in df.columns]].to_dict("records")
    # Replace NaN lat/lon with None for JSON
    for r in records:
        if "latitude" in r and (r["latitude"] is None or pd.isna(r["latitude"])):
            r["latitude"] = None
        if "longitude" in r and (r["longitude"] is None or pd.isna(r["longitude"])):
            r["longitude"] = None
    return records

# Red Sea maritime keywords for filtering regional events.
# Only very specific terms — avoids false matches from generic conflict words.
_MARITIME_KEYWORDS = [
    "houthi", "ansar allah", "red sea", "bab el-mandeb", "bab al-mandab",
    "gulf of aden", "maritime", "shipping", "vessel", "tanker", "cargo ship",
    "oil tanker", "commercial ship", "merchant vessel", "container ship",
    "suez canal", "usns", "uss ",
    "piracy", "hijack", "sea route", "waterway", "blockade",
    "coast guard", "naval blockade", "naval operation",
]


def _paginated_acled_fetch(token: str, params: dict, label: str, max_pages: int = 10) -> List[dict]:
    """Fetch paginated ACLED results.

    ACLED's multi-year paginated queries (limit=5000) routinely take 5-30s per
    page from Render's slow egress. Originally 10s → silent fallback on every
    live fetch. Bumped to 30s, then again to 60s with one retry after we still
    saw `ReadTimeout` in production — the cost of retrying once (extra 60s
    per failing page) is a one-time hit during the preload thread, but it
    keeps the dashboard off the frozen fallback when ACLED is just slow.
    """
    results = []
    for page in range(1, max_pages + 1):
        p = {**params, "page": page}
        # One retry: ACLED routinely returns ReadTimeout on the FIRST page
        # request (cold connection) but succeeds on retry.
        last_err = None
        for attempt in range(2):
            try:
                resp = requests.get(
                    config.ACLED_DATA_URL,
                    headers={**_BROWSER_HEADERS, "Authorization": f"Bearer {token}"},
                    params=p,
                    timeout=60,
                )
                resp.raise_for_status()
                last_err = None
                break
            except (requests.Timeout, requests.ConnectionError) as e:
                last_err = e
                if attempt == 0:
                    logger.warning(f"ACLED {label} page {page}: {type(e).__name__}, retrying once")
                    continue
        if last_err is not None:
            # Both attempts failed — surface the error to caller so it can
            # fall back to cache, but log clearly so we know which page died.
            raise last_err
        batch = resp.json().get("data", [])
        if not batch:
            break
        results.extend(batch)
        logger.info(f"ACLED {label}: page {page} ({len(batch)} events)")
        if len(batch) < int(params.get("limit", 5000)):
            break
    return results


def _is_maritime_relevant(event: dict) -> bool:
    """Check if an event is relevant to Red Sea / maritime operations."""
    text = f"{event.get('notes', '')} {event.get('actor1', '')} {event.get('actor2', '')}".lower()
    return any(kw in text for kw in _MARITIME_KEYWORDS)


_acled_events_memo: Optional[List[dict]] = None
_acled_events_memo_ts: float = 0.0
_acled_events_memo_lock = threading.Lock()

def fetch_acled_events() -> List[dict]:
    """Fetch comprehensive Houthi/Red Sea events from ACLED with multi-query approach.

    Strategy:
    1. All Yemen events (primary conflict zone)
    2. Houthi/Ansar Allah actor events globally (maritime attacks outside Yemen)
    3. Red Sea regional countries filtered for maritime relevance

    In-process memoization: the on-disk cache is ~13MB of JSON which
    materializes to 40-70MB of Python dicts per parse. Without the in-process
    cache, every /api/events request re-parses that blob (and on Render's
    512MB free tier two concurrent parses are enough to OOM the worker).

    Concurrency: wrapped in singleflight so N concurrent callers share a
    single fetch instead of all racing past the memo check and triggering
    parallel ACLED queries (which produced the cascade of ReadTimeouts in
    the production logs).
    """
    # Fast path BEFORE singleflight so cache hits don't even acquire the
    # singleflight lock. The actual fetch is what we want to coalesce.
    if _acled_events_memo is not None and time.time() - _acled_events_memo_ts < 600:
        return _acled_events_memo
    return _sf.run("acled_events", _do_fetch_acled_events)


def _do_fetch_acled_events() -> List[dict]:
    """Inner implementation; only one thread runs this at a time per
    singleflight semantics. The leader populates memo+cache; waiters
    receive the leader's result without re-running the fetch."""
    global _acled_events_memo, _acled_events_memo_ts
    global _acled_fetch_error, _acled_fetch_source, _acled_fetch_ts
    # Re-check the memo after the singleflight handoff in case the previous
    # leader populated it while we were waiting.
    if _acled_events_memo is not None and time.time() - _acled_events_memo_ts < 600:
        return _acled_events_memo

    cached = _read_cache("acled_events", config.CACHE_TTL_ACLED)
    if cached:
        with _acled_events_memo_lock:
            _acled_events_memo = cached
            _acled_events_memo_ts = time.time()
        _acled_fetch_source = "cache"
        _acled_fetch_ts = time.time()
        _acled_fetch_error = None
        logger.info("ACLED: serving from cache")
        return cached

    try:
        token = _get_acled_token()
        all_events = []
        seen_ids = set()

        def _add_unique(events):
            added = 0
            for e in events:
                eid = e.get("event_id_cnty")
                if eid and eid not in seen_ids:
                    all_events.append(e)
                    seen_ids.add(eid)
                    added += 1
            return added

        base_params = {
            "_format": "json",
            "event_date": _ACLED_DATE_RANGE,
            "event_date_where": "BETWEEN",
            "fields": _ACLED_FIELDS,
            "limit": 5000,
        }

        # All queries run in parallel for speed
        regional_countries = ["Saudi Arabia", "Djibouti", "Eritrea", "Oman", "Somalia", "Egypt", "Sudan", "Jordan", "Israel"]

        def _fetch_yemen():
            return "Yemen", _paginated_acled_fetch(token, {**base_params, "country": "Yemen"}, "Yemen"), False

        def _fetch_houthi():
            return "Houthi-actor", _paginated_acled_fetch(token, {
                **base_params, "actor1": "Houthi", "actor1_where": "LIKE",
            }, "Houthi-actor"), False

        def _fetch_ansar():
            return "AnsarAllah-actor", _paginated_acled_fetch(token, {
                **base_params, "actor1": "Ansar Allah", "actor1_where": "LIKE",
            }, "AnsarAllah-actor"), False

        def _fetch_regional(country):
            regional = _paginated_acled_fetch(token, {**base_params, "country": country}, country)
            maritime = [e for e in regional if _is_maritime_relevant(e)]
            return country, maritime, True

        # Cap concurrency at 3 so we only hold 3 in-flight response bodies in
        # memory at a time (previously 6 => spikes of ~80MB on Render free).
        with ThreadPoolExecutor(max_workers=3) as pool:
            futures = [
                pool.submit(_fetch_yemen),
                pool.submit(_fetch_houthi),
                pool.submit(_fetch_ansar),
            ]
            futures.extend(pool.submit(_fetch_regional, c) for c in regional_countries)

            for future in as_completed(futures):
                label, events_list, is_regional = future.result()
                n = _add_unique(events_list)
                if is_regional:
                    logger.info(f"ACLED {label}: {len(events_list)} maritime, {n} new")
                else:
                    logger.info(f"ACLED {label}: {n} unique events")

        if all_events:
            _write_cache("acled_events", all_events)
            with _acled_events_memo_lock:
                _acled_events_memo = all_events
                _acled_events_memo_ts = time.time()
            _acled_fetch_source = "api"
            _acled_fetch_ts = time.time()
            _acled_fetch_error = None
            logger.info(f"ACLED: total {len(all_events)} unique events fetched and cached")
            return all_events
        else:
            _acled_fetch_error = "ACLED API returned 0 events for every query"
            logger.warning(_acled_fetch_error)

    except Exception as e:
        _acled_fetch_error = f"{type(e).__name__}: {e}"
        logger.warning(f"ACLED API failed: {_acled_fetch_error}")

    fallback = _load_acled_fallback()
    # Memoize the fallback too — it's the same 13MB blob
    if fallback:
        with _acled_events_memo_lock:
            _acled_events_memo = fallback
            _acled_events_memo_ts = time.time()
        _acled_fetch_source = "fallback"
        _acled_fetch_ts = time.time()
    return fallback


def get_acled_fetch_error() -> Optional[str]:
    return _acled_fetch_error


def get_acled_fetch_meta() -> dict:
    """Diagnostic metadata about the most recent ACLED fetch."""
    return {
        "source": _acled_fetch_source,
        "ts": _acled_fetch_ts,
        "error": _acled_fetch_error,
        "credentials_configured": bool(config.ACLED_USERNAME and config.ACLED_PASSWORD),
    }


# ─── Chokepoint incident overlay ────────────────────────────────────────────
#
# Replaces the old AISStream.io live-vessel feed (which was unusable on Render
# free tier — single-slot key + zombie-connection cycle = perpetual 429). The
# thesis is about chokepoint *risk* and the model's dependent variable is
# event frequency, so plotting real ACLED incidents inside the chokepoint zone
# is more directly informative than vessel positions ever were.
#
# Bounding boxes are intentionally larger than the AIS kill-zone rings: ACLED
# geocodes events to launch sites or impact sites, which can be tens of km
# inland from a strait. We want to surface "Houthi missile launched from inland
# Yemen targeting a tanker in Bab" even though the launch coordinate is well
# outside the 14km kill ring.
#
# (lat_top_left, lon_top_left), (lat_bottom_right, lon_bottom_right)
INCIDENT_BOUNDING_BOXES = {
    "hormuz": ((30.5, 50.0), (22.5, 60.5)),  # Persian Gulf + Gulf of Oman + Iran coast + UAE/Oman
    "bab":    ((20.0, 39.0), (10.0, 49.0)),  # South Red Sea + Yemen + Gulf of Aden + Djibouti/Eritrea
    "suez":   ((33.0, 30.0), (27.0, 36.0)),  # Suez Canal + Sinai
}

# Actor keywords that we accept *anywhere on the globe* for each chokepoint
# (so a Houthi attack coded in inland Yemen still shows up for Bab even if
# coordinates would otherwise put it outside the bounding box). These match
# both `actor1` and `actor2` substring-insensitive.
INCIDENT_ACTOR_HINTS = {
    "hormuz": ("irgc", "iranian navy", "military forces of iran", "islamic revolutionary guard"),
    "bab":    ("houthi", "ansar allah"),
    "suez":   (),
}


def _incident_in_box(lat: float, lon: float, box) -> bool:
    (lat_n, lon_w), (lat_s, lon_e) = box
    return lat_s <= lat <= lat_n and lon_w <= lon <= lon_e


def _incident_actor_match(event: dict, hints: tuple) -> bool:
    if not hints:
        return False
    blob = ((event.get("actor1") or "") + " " + (event.get("actor2") or "")).lower()
    return any(h in blob for h in hints)


def get_chokepoint_incidents(chokepoint_id: str, days: int = 90, limit: int = 200) -> List[dict]:
    """Real ACLED + iran-events incidents inside a chokepoint zone.

    Returns events from the last `days` days that satisfy EITHER:
      (a) lat/lon falls inside the chokepoint's incident bounding box, OR
      (b) actor1/actor2 matches the chokepoint's actor hints (Houthi for Bab,
          IRGC/Iran for Hormuz) — surfaces events whose coordinates are
          inland but whose target was the chokepoint.

    Each returned record carries: event_date, event_type, sub_event_type,
    actor1, location, latitude, longitude, fatalities, source, notes.

    No synthesis. If the data files are empty or stale beyond `days`, returns
    [] and the frontend surfaces that state honestly.
    """
    if chokepoint_id not in INCIDENT_BOUNDING_BOXES:
        return []
    box = INCIDENT_BOUNDING_BOXES[chokepoint_id]
    actor_hints = INCIDENT_ACTOR_HINTS.get(chokepoint_id, ())

    # Date cutoff. We use the dataset's own newest event as the anchor (rather
    # than wall-clock now) because the cached ACLED dump can lag the calendar
    # by weeks — using wall-clock would silently produce empty results when
    # the dump is stale.

    # Prefer the LIVE ACLED memo / cache over the bundled JSON fallback. The
    # bundled file is from the last manual refresh (currently Mar 2025) and
    # was silently being preferred even when fetch_acled_events() had pulled
    # current data into the in-process memo. Result: the chokepoint incident
    # sidebar showed 13-month-old events even with a healthy ACLED API.
    acled = _acled_events_memo or []
    if not acled:
        try:
            acled = _read_cache("acled_events", config.CACHE_TTL_ACLED) or []
        except Exception:
            acled = []
    if not acled:
        acled = _load_acled_fallback() or []

    iran = _iran_fallback_memo or []
    if not iran:
        try:
            iran = _read_cache("iran_events", 86_400) or []
        except Exception:
            iran = []
    if not iran:
        iran = _load_iran_json_fallback() or []

    pool = list(acled) + list(iran)

    if not pool:
        return []

    # Find dataset anchor (newest event_date present)
    newest_ts = 0
    for e in pool:
        d = e.get("event_date") or e.get("date")
        if not d:
            continue
        try:
            ts = time.mktime(time.strptime(d[:10], "%Y-%m-%d"))
            if ts > newest_ts:
                newest_ts = ts
        except Exception:
            continue
    if newest_ts == 0:
        newest_ts = time.time()
    cutoff_ts = newest_ts - days * 86400

    seen_ids = set()
    out = []
    for e in pool:
        eid = e.get("event_id_cnty") or e.get("event_id")
        if eid:
            if eid in seen_ids:
                continue
            seen_ids.add(eid)

        # Coords (ACLED stores them as strings, sometimes empty)
        try:
            lat = float(e.get("latitude") or 0)
            lon = float(e.get("longitude") or 0)
        except (TypeError, ValueError):
            lat = lon = 0.0

        in_box = (lat or lon) and _incident_in_box(lat, lon, box)
        actor_match = _incident_actor_match(e, actor_hints)
        if not (in_box or actor_match):
            continue

        # Date filter
        d = e.get("event_date") or e.get("date") or ""
        try:
            ts = time.mktime(time.strptime(d[:10], "%Y-%m-%d"))
        except Exception:
            continue
        if ts < cutoff_ts:
            continue

        out.append({
            "event_id": eid or "",
            "date": d[:10],
            "ts": ts,
            "event_type": e.get("event_type") or "",
            "sub_event_type": e.get("sub_event_type") or "",
            "actor1": e.get("actor1") or "",
            "actor2": e.get("actor2") or "",
            "location": e.get("location") or "",
            "country": e.get("country") or "",
            "lat": round(lat, 4) if lat else None,
            "lon": round(lon, 4) if lon else None,
            "fatalities": int(e.get("fatalities") or 0),
            "source": e.get("source") or "",
            "notes": e.get("notes") or "",
            "in_box": bool(in_box),
            "actor_attributed": bool(actor_match),
        })

    # Newest first
    out.sort(key=lambda x: x["ts"], reverse=True)
    return out[:limit]


def get_chokepoint_incidents_meta() -> dict:
    """Diagnostic info for the chokepoint-incidents endpoint."""
    return {
        "source": "ACLED + iran_events JSON",
        "acled_meta": get_acled_fetch_meta(),
        "boxes": {
            cp: {"top_left": list(box[0]), "bottom_right": list(box[1])}
            for cp, box in INCIDENT_BOUNDING_BOXES.items()
        },
    }


def get_freshness_snapshot() -> dict:
    """Per-source data freshness for the UI status pill + chokepoint cards.

    Returns the newest data date and origin (api / cache / fallback) for every
    source that feeds an analytical claim. The frontend uses this to
    (a) replace the "LIVE · ACLED + EIA" pill text with an honest summary,
    (b) gate the green dot to only fire when data is genuinely fresh, and
    (c) print a "data through <date>" caption next to each chokepoint card.

    Reads memos and on-disk caches WITHOUT triggering live fetches so it's
    safe to poll on a 60s interval. The "max" date returned for each source is
    the actual newest event/observation date present, not the last fetch
    timestamp — that distinction matters when the API is reachable but the
    upstream provider hasn't updated their feed yet (common with ACLED, which
    has a documented 1-2 week lag).
    """
    import time as _time

    out: Dict[str, Dict[str, Any]] = {"server_ts": _time.time()}

    def _newest_date(events, key="event_date"):
        m = ""
        for e in events or ():
            d = (e.get(key) or e.get("date") or "")[:10]
            if d and d > m:
                m = d
        return m or None

    def _newest_field(rows, key):
        m = ""
        for r in rows or ():
            d = (r.get(key) or "")[:10]
            if d and d > m:
                m = d
        return m or None

    # ── ACLED events ─────────────────────────────────────────────────────
    # Prefer the live in-memory memo. If the memo is empty (pre-warm hasn't
    # run yet, fresh process), fall back to the on-disk cache and finally
    # to the JSON fallback file. We never trigger a network fetch from the
    # freshness endpoint — it's poll-safe.
    acled_meta = get_acled_fetch_meta()
    acled_events = _acled_events_memo or []
    acled_source_label = acled_meta.get("source")
    if not acled_events:
        try:
            disk = _read_cache("acled_events", 86_400 * 30) or []
        except Exception:
            disk = []
        if disk:
            acled_events = disk
            if not acled_source_label:
                acled_source_label = "cache"
    if not acled_events:
        try:
            fb_path = config.DATA_DIR / "acled_events.json"
            if fb_path.exists():
                acled_events = json.loads(fb_path.read_text())
                if not acled_source_label:
                    acled_source_label = "fallback"
        except Exception:
            pass
    out["acled"] = {
        "newest_date": _newest_date(acled_events),
        "count": len(acled_events),
        "source": acled_source_label,                # "api" | "cache" | "fallback"
        "fetched_ts": acled_meta.get("ts"),
        "error": acled_meta.get("error"),
    }

    # ── Iran events ──────────────────────────────────────────────────────
    iran_events = _iran_fallback_memo or []
    if not iran_events:
        # Probe the on-disk cache without forcing a live fetch
        try:
            iran_events = _read_cache("iran_events", 86_400 * 30) or []
        except Exception:
            iran_events = []
    if not iran_events:
        # Final fallback: bundled iran_events.json (same source the data
        # service uses when the live API is unreachable).
        try:
            iran_events = _load_iran_json_fallback() or []
        except Exception:
            iran_events = []
    out["iran"] = {
        "newest_date": _newest_date(iran_events),
        "count": len(iran_events),
        "error": get_iran_fetch_error(),
    }

    # ── Brent (cache → CSV fallback) ─────────────────────────────────────
    try:
        brent_cached = _read_cache("brent_prices", 86_400 * 30) or []
    except Exception:
        brent_cached = []
    if not brent_cached:
        try:
            brent_cached = _load_brent_csv_fallback() or []
        except Exception:
            brent_cached = []
    out["brent"] = {
        "newest_date": _newest_field(brent_cached, "date"),
        "count": len(brent_cached),
    }

    # ── DXY / OVX (yfinance) ─────────────────────────────────────────────
    for k, cache_key in (("dxy", "dxy"), ("ovx", "ovx")):
        try:
            rows = _read_cache(cache_key, 86_400 * 7) or []
        except Exception:
            rows = []
        out[k] = {
            "newest_date": _newest_field(rows, "date"),
            "count": len(rows),
        }

    # ── PortWatch transit data (monthly) ─────────────────────────────────
    for k, fn_name, cache_key in (
        ("hormuz", "fetch_hormuz_transits", "hormuz_transits"),
        ("bab",    "fetch_bab_el_mandeb_transits", "bab_el_mandeb_transits"),
        ("suez",   "fetch_suez_transits", "suez_transits"),
    ):
        try:
            rows = _read_cache(cache_key, 86_400 * 30) or []
        except Exception:
            rows = []
        # PortWatch returns monthly buckets keyed by `month`
        latest_month = ""
        for r in rows:
            m = (r.get("month") or "")[:7]
            if m and m > latest_month:
                latest_month = m
        out[k + "_transits"] = {
            "newest_month": latest_month or None,
            "count": len(rows),
        }

    # ── HDX live ACLED mirror (no-auth, no live fetch — just read cache) ─
    # If the HDX cache exists, surface its freshness so the pill upgrades
    # from "frozen" → "live" once the user has called /api/live-event-counts
    # at least once. Never trigger a network fetch from here — keep this
    # endpoint poll-safe.
    hdx_newest_month = None
    hdx_country_max: Dict[str, str] = {}
    hdx_fetched_utc = None
    try:
        hdx_cached = _read_cache("hdx_event_counts", 86_400 * 7) or {}
        hdx_newest_month = hdx_cached.get("newest_month")
        hdx_fetched_utc  = hdx_cached.get("fetched_utc")
        for country, rows in (hdx_cached.get("by_country") or {}).items():
            if rows:
                hdx_country_max[country] = rows[-1].get("month")
    except Exception:
        pass
    out["hdx_acled"] = {
        "newest_month": hdx_newest_month,
        "by_country_newest_month": hdx_country_max,
        "fetched_utc": hdx_fetched_utc,
    }

    # ── Aggregate health: green if every critical source has data within
    #    30 days of its respective newest record; amber if 30-180d; red if
    #    older or missing. The frontend uses this single field to color
    #    the top-right status pill. HDX freshness is preferred over ACLED
    #    fallback when available (HDX is updated weekly; the bundled JSON
    #    can be months stale).
    today = _time.strftime("%Y-%m-%d", _time.gmtime(_time.time()))

    def _days_old(d):
        if not d:
            return 10_000
        try:
            t = _time.mktime(_time.strptime(d[:10], "%Y-%m-%d"))
            return max(0, int((_time.time() - t) / 86_400))
        except Exception:
            return 10_000

    # Effective ACLED freshness: prefer the HDX monthly aggregate (which is
    # always within ~7 days of wall-clock when reachable) over the bundled
    # JSON fallback (which can be 13+ months old). HDX gives us "YYYY-MM";
    # treat the 1st of the following month as the freshness boundary.
    def _month_end(ym):
        if not ym or len(ym) < 7:
            return None
        try:
            y = int(ym[:4]); m = int(ym[5:7])
            # End-of-month estimate so a 2026-04 month doesn't read as
            # "2026-04-01" (1st) which would be artificially stale.
            ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
            return f"{ny:04d}-{nm:02d}-01"
        except Exception:
            return None
    hdx_eff_date = _month_end(out["hdx_acled"]["newest_month"])
    acled_eff_date = out["acled"]["newest_date"]
    if hdx_eff_date and (not acled_eff_date or hdx_eff_date > acled_eff_date):
        acled_eff_date = hdx_eff_date
        out["acled"]["effective_source"] = "hdx_mirror"
    else:
        out["acled"]["effective_source"] = out["acled"].get("source") or "bundled"
    out["acled"]["effective_date"] = acled_eff_date

    critical_age = max(
        _days_old(acled_eff_date),
        _days_old(out["brent"]["newest_date"]),
    )
    if critical_age <= 7:
        status = "live"
    elif critical_age <= 45:
        status = "stale"
    else:
        status = "frozen"
    out["status"] = {
        "level": status,                # "live" | "stale" | "frozen"
        "critical_age_days": critical_age,
        "today_utc": today,
    }
    return out


def _load_acled_fallback() -> List[dict]:
    """Load ACLED data from JSON fallback, then CSV files."""
    # Try JSON fallback first (pre-fetched comprehensive dataset)
    json_path = config.DATA_DIR / "acled_events.json"
    if json_path.exists():
        try:
            events = json.loads(json_path.read_text())
            logger.info(f"ACLED fallback: loaded {len(events)} events from acled_events.json")
            return events
        except Exception as e:
            logger.warning(f"ACLED JSON fallback failed: {e}")

    # Fall back to thesis events CSV
    logger.info("ACLED: falling back to thesis events CSV")
    if config.THESIS_EVENTS_PATH.exists():
        try:
            df = pd.read_csv(config.THESIS_EVENTS_PATH)
            col_map = {c: c.lower() for c in df.columns}
            df.rename(columns=col_map, inplace=True)
            records = _df_to_event_records(df)
            logger.info(f"ACLED CSV fallback: loaded {len(records)} events from {config.THESIS_EVENTS_PATH.name}")
            return records
        except Exception as e:
            logger.warning(f"Failed to load thesis events CSV: {e}")
    return []


# ─── Thesis Dataset (726 Verified Maritime Events) ───────────────────────────

_thesis_events_cache: Optional[List[dict]] = None

def load_thesis_events() -> List[dict]:
    """Load the 726 ACLED-verified maritime events analyzed in the thesis."""
    global _thesis_events_cache
    if _thesis_events_cache is not None:
        return _thesis_events_cache

    csv_path = config.THESIS_EVENTS_PATH
    if not csv_path.exists():
        logger.warning("No thesis events CSV found, falling back to ACLED")
        return fetch_acled_events()

    try:
        df = pd.read_csv(csv_path).copy()
        col_map = {c: c.lower() for c in df.columns}
        df = df.rename(columns=col_map)

        # Map CH6 column names to expected schema
        if "date" in df.columns and "event_date" not in df.columns:
            df = df.rename(columns={"date": "event_date"})
        if "event_id" in df.columns and "event_id_cnty" not in df.columns:
            df = df.rename(columns={"event_id": "event_id_cnty"})

        records = _df_to_event_records(df.copy())
        _thesis_events_cache = records
        logger.info(f"Thesis dataset: loaded {len(records)} events from {csv_path.name}")
        return records
    except Exception as e:
        logger.warning(f"Failed to load thesis events: {e}")
        return fetch_acled_events()


# ─── EIA API v2 (Brent Crude + SPR) ─────────────────────────────────────────

def fetch_brent_prices() -> List[dict]:
    """Fetch daily Brent crude spot prices from EIA API v2."""
    cached = _read_cache("brent_prices", config.CACHE_TTL_BRENT)
    if cached:
        logger.info("Brent: serving from cache")
        return cached

    try:
        resp = requests.get(
            f"{config.EIA_BASE_URL}/petroleum/pri/spt/data",
            params={
                "api_key": config.EIA_API_KEY,
                "frequency": "daily",
                "data[0]": "value",
                "facets[series][]": "RBRTE",
                "sort[0][column]": "period",
                "sort[0][direction]": "asc",
                "start": "2023-10-01",
                "length": 5000,
            },
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        records = [
            {"date": r["period"], "price": float(r["value"])}
            for r in data.get("response", {}).get("data", [])
            if r.get("value") is not None
        ]
        if records:
            records = _supplement_brent_recent(records)
            _write_cache("brent_prices", records)
            logger.info(f"Brent: fetched {len(records)} daily prices (EIA + yfinance)")
            return records
    except Exception as e:
        logger.warning(f"EIA Brent API failed: {e}")

    fallback = _load_brent_csv_fallback()
    if fallback:
        fallback = _supplement_brent_recent(fallback)
    return fallback


def _supplement_brent_recent(eia_records: List[dict]) -> List[dict]:
    """Extend EIA/FRED Brent data with recent prices beyond API reporting lag.

    Data waterfall (each layer only fills gaps left by the one above):
        1. Yahoo Finance direct API  (real-time, but often blocked from servers)
        2. yfinance library           (same data, different client)
        3. FRED DCOILBRENTEU series   (reliable, ~1-day lag, auto-updates daily)
        4. Hardcoded fallback table   (last resort for known war-period dates)
    """
    # Deduplicate by date (keep latest) and sort chronologically
    by_date = {r["date"]: r for r in eia_records}
    last_date = max(by_date.keys()) if by_date else "2023-10-01"

    # Hardcoded war-period settlements from CNBC / Reuters / Bloomberg.
    # Used ONLY when ALL live sources (EIA, Yahoo, FRED) fail for a given date.
    fallback_prices = {
        "2026-03-02": 77.24,   # First trading day after war started Feb 28 (Sat)
        "2026-03-03": 83.28,   # Brent surges as shipping suspended
        "2026-03-04": 81.56,   # Slight pullback amid heavy strikes
        "2026-03-05": 88.59,   # Brent surges on insurance withdrawal, 500+ missiles
        "2026-03-06": 95.74,   # Analysts warn $100+; Iran strikes Gulf states
        # Mar 7 (Sat) and Mar 8 (Sun) — no settlement
        "2026-03-09": 98.96,   # Brent settles +3.4% from Friday
        "2026-03-10": 87.80,   # Sharp pullback (-11.3%) as Trump signals war "very complete"
        "2026-03-11": 91.98,   # CNBC: Brent +4.76% as IEA releases 400M bbl reserves
        "2026-03-12": 100.46,  # CNBC: First close above $100 since Aug 2022 (+9.22%)
        "2026-03-13": 103.14,  # CNBC: Brent +2.67%; oil above $100 for second day
        # Mar 14 (Sat) and Mar 15 (Sun) — no settlement
        "2026-03-16": 101.04,  # EIA/FRED spot price
        "2026-03-17": 103.42,  # CNBC: +3.2% as allies refuse Hormuz escort
        "2026-03-18": 107.38,  # CNBC: +3.8% amid Hormuz shutdown fears
        "2026-03-19": 108.65,  # CNBC: Hit $119 intraday; settled +1.18% after Netanyahu comments
        "2026-03-20": 110.96,  # Iraq force majeure on Basra crude; supply disruptions widen
        # Mar 21 (Sat) and Mar 22 (Sun) — no settlement
        "2026-03-23": 96.07,   # Sharp pullback on IRGC Hormuz 'tollbooth' de-escalation signals
        "2026-03-24": 100.09,  # Partial recovery; China/India negotiate Hormuz passage
        "2026-03-25": 103.42,  # Kuwait airport drone attack; regional spread
        "2026-03-26": 108.01,  # Iran rejects talks; IRGC confirms tollbooth system
        "2026-03-27": 105.32,  # Marines arrive; humanitarian corridor eases fears slightly
        "2026-03-28": 112.57,  # Houthis fire missiles at Israel; 82nd Airborne deploys
        # Mar 29 (Sat) and Mar 30 (Sun) — no settlement
        "2026-03-31": 115.24,  # Continued airstrikes; parliament rejects negotiations
        "2026-04-01": 113.88,  # UK 35-nation Hormuz meeting; diplomatic push
        "2026-04-02": 111.45,  # Trump says objectives nearly met; coalition diplomacy
        "2026-04-03": 118.32,  # US F-15E shot down over Iran; escalation fears
        "2026-04-04": 121.07,  # Bushehr nuclear plant struck; IAEA warning
        # Apr 5 (Sat) and Apr 6 (Sun) — no settlement
        "2026-04-07": 126.00,  # Trump 'civilization will die tonight'; ceasefire announced
        "2026-04-08": 93.76,   # Ceasefire crash — Brent drops ~13% from highs
        "2026-04-09": 100.99,  # Rebound as Hormuz stays closed despite ceasefire
        "2026-04-10": 97.78,   # Trump accuses Iran of poor Hormuz management
        "2026-04-11": 95.71,   # US Navy enters Hormuz; Islamabad talks begin
        # Apr 12 (Sat) — no settlement
        "2026-04-13": 103.72,  # Surges 6.95% on US naval blockade announcement
        "2026-04-14": 100.19,  # Blockade takes effect; Trump hints at resumed talks
        "2026-04-15": 96.83,   # Diplomatic optimism — "very close to over"
        "2026-04-16": 94.89,   # Lebanon 10-day truce announced; regional tensions ease
        "2026-04-17": 93.50,   # Hormuz declared open during ceasefire — relief continues
    }

    # ── Layers 1-3: Run ALL live sources in parallel ──────────────────────────
    # Yahoo Finance almost always fails from Render cloud IPs, so don't wait
    # sequentially. FRED is the most reliable server-side source.

    def _try_yahoo_direct():
        """Layer 1: Yahoo Finance direct API."""
        try:
            last_dt = datetime.strptime(last_date, "%Y-%m-%d")
            period1 = int(last_dt.timestamp())
            period2 = int((datetime.now() + timedelta(days=1)).timestamp())
            yf_url = (
                f"https://query1.finance.yahoo.com/v8/finance/chart/BZ=F"
                f"?period1={period1}&period2={period2}&interval=1d"
            )
            yf_resp = requests.get(yf_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=3)
            yf_resp.raise_for_status()
            yf_data = yf_resp.json()
            result = yf_data.get("chart", {}).get("result", [])
            prices = {}
            if result:
                timestamps = result[0].get("timestamp", [])
                closes = result[0]["indicators"]["quote"][0].get("close", [])
                for ts, close in zip(timestamps, closes):
                    if close is None:
                        continue
                    d = datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
                    if d > last_date:
                        prices[d] = {"date": d, "price": round(float(close), 2)}
            return prices
        except Exception as e:
            logger.warning(f"Yahoo Finance API failed: {e}")
            return {}

    def _try_yfinance_lib():
        """Layer 2: yfinance library."""
        try:
            import yfinance as yf
            df = yf.download("BZ=F", start=last_date, progress=False)
            if df.empty:
                return {}
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            df_valid = df[df["Close"].notna()].copy()
            df_valid["_date"] = df_valid.index.strftime("%Y-%m-%d")
            df_valid = df_valid[df_valid["_date"] > last_date]
            return {d: {"date": d, "price": round(float(p), 2)} for d, p in zip(df_valid["_date"], df_valid["Close"])}
        except Exception as e:
            logger.warning(f"Brent yfinance failed: {e}")
            return {}

    def _try_fred():
        """Layer 3: FRED DCOILBRENTEU (most reliable, ~1-day lag)."""
        try:
            from fredapi import Fred
            fred = Fred(api_key=config.FRED_API_KEY)
            series = fred.get_series("DCOILBRENTEU", observation_start=last_date)
            prices = {}
            for idx, val in series.items():
                if pd.notna(val):
                    d = idx.strftime("%Y-%m-%d")
                    if d > last_date:
                        prices[d] = {"date": d, "price": round(float(val), 2)}
            return prices
        except Exception as e:
            logger.warning(f"FRED Brent (DCOILBRENTEU) failed: {e}")
            return {}

    # Run all 3 sources in parallel — first to return wins for each date
    with ThreadPoolExecutor(max_workers=3) as pool:
        f_yahoo = pool.submit(_try_yahoo_direct)
        f_yf = pool.submit(_try_yfinance_lib)
        f_fred = pool.submit(_try_fred)

    # Merge results: FRED first (most reliable), then overlay Yahoo/yfinance
    for source_prices in [f_fred.result(), f_yf.result(), f_yahoo.result()]:
        for d, rec in source_prices.items():
            if d not in by_date:
                by_date[d] = rec

    live_added = sum(1 for d in by_date if d > last_date)
    if live_added:
        logger.info(f"Brent: {live_added} live prices supplemented after {last_date}")

    # ── Layer 4: Hardcoded fallback (last resort) ────────────────────────────
    for date_str, price in fallback_prices.items():
        if date_str not in by_date:
            by_date[date_str] = {"date": date_str, "price": price}
            logger.info(f"Brent: using fallback price for {date_str}: ${price}")
    if len(by_date) > len(eia_records):
        logger.info(f"Brent: supplemented {len(by_date) - len(eia_records)} prices after {last_date}")

    return sorted(by_date.values(), key=lambda r: r["date"])


def _load_brent_csv_fallback() -> List[dict]:
    logger.info("Brent: falling back to CSV")
    if config.MYLES_DATASET_PATH.exists():
        df = pd.read_csv(config.MYLES_DATASET_PATH)
        date_col = df.columns[0]
        df.rename(columns={date_col: "Date"}, inplace=True)
        df["Date"] = pd.to_datetime(df["Date"])
        df_valid = df[df["Brent_Price"].notna()].copy()
        records = [
            {"date": d.strftime("%Y-%m-%d"), "price": float(p)}
            for d, p in zip(df_valid["Date"], df_valid["Brent_Price"])
        ]
        return records
    return []


def fetch_spr_data() -> List[dict]:
    """Fetch SPR stock levels from EIA API v2."""
    cached = _read_cache("spr_data", config.CACHE_TTL_BRENT)
    if cached:
        return cached

    try:
        resp = requests.get(
            f"{config.EIA_BASE_URL}/petroleum/stoc/wstk/data",
            params={
                "api_key": config.EIA_API_KEY,
                "frequency": "weekly",
                "data[0]": "value",
                "facets[product][]": "EPC0",
                "facets[process][]": "SAX",
                "sort[0][column]": "period",
                "sort[0][direction]": "asc",
                "start": "2023-10-01",
                "length": 5000,
            },
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        records = [
            {"date": r["period"], "value": float(r["value"])}
            for r in data.get("response", {}).get("data", [])
            if r.get("value") is not None
        ]
        if records:
            _write_cache("spr_data", records)
            logger.info(f"SPR: fetched {len(records)} weekly data points")
            return records
    except Exception as e:
        logger.warning(f"EIA SPR API failed: {e}")
    return []


# ─── yfinance (DXY, OVX) ────────────────────────────────────────────────────

def fetch_yfinance_series(ticker: str, cache_key: str) -> List[dict]:
    """Fetch daily time series from Yahoo Finance."""
    cached = _read_cache(cache_key, config.CACHE_TTL_YFINANCE)
    if cached:
        logger.info(f"yfinance {ticker}: serving from cache")
        return cached

    try:
        import yfinance as yf
        df = yf.download(ticker, start="2023-10-01", progress=False)
        if df.empty:
            raise ValueError(f"No data returned for {ticker}")

        # Handle multi-level columns from yfinance
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        df_valid = df[df["Close"].notna()].copy()
        records = [
            {"date": idx.strftime("%Y-%m-%d"), "value": round(float(close), 4)}
            for idx, close in zip(df_valid.index, df_valid["Close"])
        ]
        if records:
            _write_cache(cache_key, records)
            logger.info(f"yfinance {ticker}: fetched {len(records)} data points")
            return records
    except Exception as e:
        logger.warning(f"yfinance {ticker} failed: {e}")
    return []


def _fred_fallback(series_id: str, cache_key: str, label: str) -> List[dict]:
    """Shared FRED fetch used as the primary source for DXY + OVX on cloud
    deploys where yfinance is blocked by Yahoo's bot detection. Writes to
    the same cache key the yfinance path used, so downstream readers don't
    care about the source change."""
    cached = _read_cache(cache_key, config.CACHE_TTL_YFINANCE)
    if cached:
        return cached
    try:
        from fredapi import Fred
        fred = Fred(api_key=config.FRED_API_KEY)
        series = fred.get_series(series_id, observation_start="2023-10-01")
        records = [
            {"date": idx.strftime("%Y-%m-%d"), "value": round(float(val), 4)}
            for idx, val in series.items() if pd.notna(val)
        ]
        if records:
            _write_cache(cache_key, records)
            logger.info(f"FRED {series_id} ({label}): fetched {len(records)} data points")
            return records
        logger.warning(f"FRED {series_id} ({label}) returned 0 rows")
    except Exception as e:
        logger.warning(f"FRED {series_id} ({label}) failed: {e}")
    stale = _read_stale_cache(cache_key)
    if stale:
        logger.info(f"FRED {series_id} ({label}): serving stale cache ({len(stale)} rows)")
        return stale
    return []


def fetch_dxy() -> List[dict]:
    """DXY (US Dollar Index).

    FRED's DTWEXBGS (broad trade-weighted dollar) is the primary source.
    yfinance's DX-Y.NYB is tried only as a secondary check — on any cloud IP
    (Render, AWS, GCP) Yahoo blocks us with an empty/HTML body, so hitting
    yfinance first wasted cycles and spammed the deploy log. The FRED series
    is slightly different math but tracks DXY within 1-2% and is how every
    thesis-grade macro dashboard sources this number anyway.
    """
    result = _fred_fallback("DTWEXBGS", "dxy", "DXY proxy")
    if result:
        return result
    # Last-ditch yfinance attempt for local dev where Yahoo still works
    return fetch_yfinance_series("DX-Y.NYB", "dxy")


def fetch_ovx() -> List[dict]:
    """OVX (CBOE Crude Oil ETF Volatility Index).

    FRED's OVXCLS is the same series CBOE publishes — FRED is the primary
    source on cloud. yfinance ^OVX is tried only as a local-dev convenience.
    """
    result = _fred_fallback("OVXCLS", "ovx", "OVX")
    if result:
        return result
    return fetch_yfinance_series("^OVX", "ovx")


# ─── FRED API (China BCI) ───────────────────────────────────────────────────

def fetch_china_pmi() -> List[dict]:
    """Fetch China business confidence (FRED BSCICP03CNM665S, monthly).

    KNOWN ISSUE (verified 2026-04-26): FRED has stopped publishing
    BSCICP03CNM665S past Jan 2024 — `fred.get_series()` returns 0 points for
    any observation_start later than that. The frozen Jan 2024 reading is
    carried forward into the master timeseries via _extend_master_timeseries
    so the regression's china_pmi control variable stays defined for new
    rows; downstream charts handle the flat-line. If/when FRED republishes
    or we adopt a substitute (CSCICP02CNM665S / Caixin PMI / NBS feed) this
    fetcher will start returning fresh points automatically.

    Stale-cache fallback so an FRED outage doesn't blank the macro panel.
    """
    cached = _read_cache("china_pmi", config.CACHE_TTL_FRED)
    if cached:
        return cached

    try:
        from fredapi import Fred
        fred = Fred(api_key=config.FRED_API_KEY)
        series = fred.get_series("BSCICP03CNM665S", observation_start="2023-10-01")
        records = [
            {"date": idx.strftime("%Y-%m-%d"), "value": round(float(val), 4)}
            for idx, val in series.items()
            if pd.notna(val)
        ]
        if records:
            _write_cache("china_pmi", records)
            logger.info(f"FRED China BCI: fetched {len(records)} data points")
            return records
    except Exception as e:
        logger.warning(f"FRED China PMI fetch failed: {e}")

    stale = _read_stale_cache("china_pmi") or []
    if stale:
        logger.info(f"FRED China PMI: serving stale cache ({len(stale)} pts)")
    return stale


# ─── Tier-1 Free API Integrations (EIA inventories, FRED macro, GDELT) ──────
# These endpoints enrich the Overview and US-Iran tabs with context that
# exec-level readers expect on an oil-price dashboard: physical supply
# (crude inventories + Cushing), forward inflation (5y breakeven), broad
# risk (VIX, HY credit spreads), and media attention (GDELT tone timeline).

def _read_stale_cache(cache_key: str):
    """Read a cache file ignoring TTL — used as a fallback when a live fetch fails.

    Better to show yesterday's data than to show nothing while the upstream
    provider has a hiccup. The dashboard has a staleness banner elsewhere.
    """
    path = _cache_path(cache_key)
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text())
        return data.get("payload")
    except Exception:
        return None


def _eia_weekly_stocks(series_id: str, cache_key: str) -> List[dict]:
    """Internal helper — fetch a weekly stocks series from EIA v2 by series ID."""
    cached = _read_cache(cache_key, config.CACHE_TTL_BRENT)
    if cached:
        return cached
    try:
        resp = requests.get(
            f"{config.EIA_BASE_URL}/petroleum/stoc/wstk/data",
            params={
                "api_key": config.EIA_API_KEY,
                "frequency": "weekly",
                "data[0]": "value",
                "facets[series][]": series_id,
                "sort[0][column]": "period",
                "sort[0][direction]": "asc",
                "start": "2023-10-01",
                "length": 5000,
            },
            timeout=30,
            headers=_BROWSER_HEADERS,
        )
        resp.raise_for_status()
        data = resp.json()
        rows = data.get("response", {}).get("data", [])
        records = [
            {"date": r["period"], "value": float(r["value"])}
            for r in rows
            if r.get("value") is not None
        ]
        if records:
            _write_cache(cache_key, records)
            logger.info(f"EIA {series_id}: fetched {len(records)} weekly data points")
            return records
        # Empty response — log the full payload so we can see WHY (e.g. auth error body)
        logger.warning(f"EIA {series_id} returned 0 rows. response={str(data)[:400]}")
    except Exception as e:
        logger.warning(f"EIA {series_id} failed: {e}")
    # Fallback: stale cache beats an empty response in the UI
    stale = _read_stale_cache(cache_key)
    if stale:
        logger.info(f"EIA {series_id}: serving stale cache ({len(stale)} rows)")
        return stale
    return []


def fetch_eia_inventories() -> Dict[str, List[dict]]:
    """Fetch weekly US commercial crude inventories + Cushing hub stocks.

    Critical context for an oil dashboard: draws on inventories during the war
    quantify physical supply tightness that price alone can obscure. Cushing
    (WTI delivery hub) matters because low stocks there amplify backwardation.
    """
    commercial = _eia_weekly_stocks("WCESTUS1", "eia_commercial_crude")
    cushing = _eia_weekly_stocks("W_EPC0_SAX_YCUOK_MBBL", "eia_cushing_stocks")
    return {"commercial_crude": commercial, "cushing": cushing}


def _fred_series(series_id: str, cache_key: str, ttl: int = None) -> List[dict]:
    """Internal helper — fetch a FRED series with cache.

    On live-fetch failure or empty response, falls back to the stale cache file
    so the UI keeps showing yesterday's numbers instead of going blank. Same
    rationale as `_eia_weekly_stocks`: freshness matters less than continuity
    for an exec dashboard.
    """
    cached = _read_cache(cache_key, ttl if ttl is not None else config.CACHE_TTL_YFINANCE)
    if cached:
        return cached
    try:
        from fredapi import Fred
        fred = Fred(api_key=config.FRED_API_KEY)
        series = fred.get_series(series_id, observation_start="2023-10-01")
        records = [
            {"date": idx.strftime("%Y-%m-%d"), "value": round(float(val), 4)}
            for idx, val in series.items()
            if pd.notna(val)
        ]
        if records:
            _write_cache(cache_key, records)
            logger.info(f"FRED {series_id}: fetched {len(records)} data points")
            return records
        logger.warning(f"FRED {series_id} returned 0 rows")
    except Exception as e:
        logger.warning(f"FRED {series_id} failed: {e}")
    # Fallback: stale cache beats an empty response in the UI
    stale = _read_stale_cache(cache_key)
    if stale:
        logger.info(f"FRED {series_id}: serving stale cache ({len(stale)} rows)")
        return stale
    return []


def fetch_macro_context() -> Dict[str, List[dict]]:
    """Fetch macro-financial context series from FRED.

    - DCOILWTICO: WTI crude spot (for Brent-WTI spread — key freight / Cushing signal)
    - T5YIE: 5-year breakeven inflation (market's pricing of war→inflation pass-through)
    - VIXCLS: S&P 500 volatility (broad risk-on/risk-off gauge)
    - BAMLH0A0HYM2EY: US High-Yield effective yield (credit spread stress)
    """
    return {
        "wti":        _fred_series("DCOILWTICO",       "fred_wti"),
        "breakeven5": _fred_series("T5YIE",            "fred_breakeven5"),
        "vix":        _fred_series("VIXCLS",           "fred_vix"),
        "hy_yield":   _fred_series("BAMLH0A0HYM2EY",   "fred_hy_yield"),
    }


def fetch_gdelt_tone(timespan: str = "90d") -> dict:
    """Fetch GDELT DOC 2.0 TimelineTone for Iran-related coverage.

    Keyless public API. Returns a daily timeline of average tone (–100 very
    negative, +100 very positive) for English-language news mentioning Iran
    in an oil/strait/military context. A sharp negative swing typically
    precedes or accompanies market stress.

    Also fetches TimelineVolRaw (raw article volume) so the frontend can
    overlay attention intensity alongside tone.

    Singleflight-wrapped: GDELT throttles aggressively (HTTP 429) when
    multiple parallel callers hit it. Coalescing concurrent requests for
    the same timespan lets one fetch service all waiters.
    """
    cached = _read_cache(f"gdelt_tone_{timespan}", 3600)  # 1-hour TTL — news-cycle data
    if cached:
        return cached
    return _sf.run(f"gdelt_tone_{timespan}", lambda: _do_fetch_gdelt_tone(timespan))


def _do_fetch_gdelt_tone(timespan: str) -> dict:
    # Re-check cache after singleflight handoff (the leader may have
    # populated it while we were waiting).
    cached = _read_cache(f"gdelt_tone_{timespan}", 3600)
    if cached:
        return cached

    query = 'iran AND (oil OR "strait of hormuz" OR military OR strikes OR sanctions)'
    base = "https://api.gdeltproject.org/api/v2/doc/doc"
    result = {"tone": [], "volume": []}

    for mode, key in [("TimelineTone", "tone"), ("TimelineVolRaw", "volume")]:
        try:
            resp = requests.get(
                base,
                params={
                    "mode": mode,
                    "query": query,
                    "format": "json",
                    "timespan": timespan,
                    "sort": "datedesc",
                },
                headers=_BROWSER_HEADERS,
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            timeline = data.get("timeline", [])
            if timeline and timeline[0].get("data"):
                # GDELT returns ISO dates like "20260415T000000Z"
                records = []
                for r in timeline[0]["data"]:
                    raw_date = r.get("date", "")
                    if len(raw_date) >= 8:
                        date_str = f"{raw_date[0:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
                        val = r.get("value")
                        if val is not None:
                            records.append({"date": date_str, "value": round(float(val), 4)})
                result[key] = records
                logger.info(f"GDELT {mode}: fetched {len(records)} daily points")
        except Exception as e:
            logger.warning(f"GDELT {mode} failed: {e}")

    # Only cache when we actually got data. Caching an empty payload would
    # block the live retry path for the full TTL window — exactly what
    # caused the §09b chart to show "no tone series" on the HF Space's
    # cold-start fetch (GDELT slow on first hit, empty response cached for
    # an hour, frontend showed empty state until the user noticed).
    if result["tone"] or result["volume"]:
        _write_cache(f"gdelt_tone_{timespan}", result)
        return result
    # Both branches empty — fall back to stale cache rather than blank chart.
    # Don't write the empty result; next request will retry the live API.
    stale = _read_stale_cache(f"gdelt_tone_{timespan}")
    if stale:
        logger.info(f"GDELT: serving stale cache (tone={len(stale.get('tone') or [])})")
        return stale
    return result


# ─── HDX Live ACLED Mirror (no-auth, weekly refresh) ────────────────────────
#
# Why this exists: the bundled ACLED JSON fallback in data/acled_events.json is
# point-in-time (as of the last manual refresh). When the live ACLED OAuth API
# is unreachable (Render egress timeouts, expired credentials, etc.) the
# dashboard silently freezes at the fallback's date.
#
# ACLED publishes country-level monthly aggregates to the Humanitarian Data
# Exchange (data.humdata.org) every Wednesday — typically 4–7 days behind
# wall-clock. Those aggregates are fetched without authentication and so
# provide a guaranteed-fresh "events per month" series even when the
# authenticated API is down. The data is granular enough to populate
# (a) a freshness timestamp visible in the UI status pill, and
# (b) per-country monthly event counts shown next to chokepoint cards.
#
# We surface this through /api/live-event-counts and merge the freshness
# timestamp into get_freshness_snapshot() so the pill says
# "EVENTS THROUGH <month> · ACLED HDX" instead of the frozen fallback date.

# Known package IDs on HDX. Each ACLED country dataset on HDX has a stable
# slug — we hit the package_show metadata to find the active resource URL
# (which rotates weekly when ACLED republishes).
_HDX_API = "https://data.humdata.org/api/3/action/package_show"
_HDX_PACKAGES = {
    "yemen": "yemen-acled-conflict-data",
    "iran":  "iran-acled-conflict-data",
    "saudi": "saudi-arabia-acled-conflict-data",
    "egypt": "egypt-acled-conflict-data",
    "lebanon": "lebanon-acled-conflict-data",
}


def _hdx_resource_url(package_id: str) -> Optional[Dict[str, Any]]:
    """Look up the most recent 'political_violence' XLSX URL for an HDX package.

    Returns {url, name, last_modified} or None on failure.
    """
    try:
        resp = requests.get(_HDX_API, params={"id": package_id},
                            headers=_BROWSER_HEADERS, timeout=15)
        if not resp.ok:
            return None
        result = resp.json().get("result", {}) or {}
        for r in result.get("resources", []) or []:
            name = (r.get("name") or "").lower()
            if "political_violence" in name and (r.get("format", "").lower() == "xlsx"):
                return {
                    "url": r.get("url"),
                    "name": r.get("name"),
                    "last_modified": (r.get("last_modified") or result.get("metadata_modified") or "")[:19],
                }
    except Exception as e:
        logger.warning(f"HDX lookup failed for {package_id}: {e}")
    return None


_HDX_MONTH_LOOKUP = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7, "aug": 8,
    "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_hdx_monthly_xlsx(content: bytes, country_label: str) -> List[dict]:
    """Parse one HDX ACLED admin-level XLSX into [{month, events, fatalities}].

    LOW-MEMORY STREAMING IMPLEMENTATION. The previous version used
    `pd.read_excel(...)` which loads the entire ~45k-row sheet into a
    DataFrame plus several rename/groupby intermediate copies — peak
    ~30MB resident per country, ~150MB across all five. On Render's
    512MB free tier that was enough (combined with ACLED's 70MB blob)
    to OOM the worker mid-preload.

    This rewrite uses openpyxl in `read_only=True` mode and streams rows
    one at a time, accumulating `(year, month) → (events, fatalities)`
    in a plain dict. No DataFrame is ever materialised. Per-country
    peak memory drops to ~3–5MB.

    HDX's ACLED files are workbooks with two sheets:
      * "TOU"   — terms-of-use boilerplate
      * "Data"  — admin-level rows: Country, Admin1, Admin2, ISO3,
                  Admin2 Pcode, Admin1 Pcode, Month (English name),
                  Year (int), Events, Fatalities
    """
    import io as _io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"HDX XLSX open failed for {country_label}: {e}")
        return []

    try:
        # Pick the data sheet by name (fallback: any sheet that isn't TOU).
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() == "data":
                sheet_name = name
                break
        if sheet_name is None:
            for name in wb.sheetnames:
                if name.lower() != "tou":
                    sheet_name = name
                    break
        if sheet_name is None:
            logger.warning(f"HDX {country_label}: no data sheet found")
            return []
        ws = wb[sheet_name]

        # Header row → column index map. Lower-cased + substring matching
        # keeps the parser robust to ACLED relabelling columns.
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        col_idx: Dict[str, int] = {}
        for i, cell in enumerate(header):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            col_idx[key] = i

        def _col(*needles):
            for k, i in col_idx.items():
                if all(n in k for n in needles):
                    return i
            return None

        year_i  = _col("year")
        month_i = _col("month")
        ev_i    = _col("event") or _col("incident") or _col("count")
        fat_i   = _col("fatal")
        if year_i is None or month_i is None or ev_i is None:
            logger.warning(f"HDX {country_label}: unrecognised schema {list(col_idx.keys())[:8]}")
            return []

        # Stream rows, accumulating into a small dict. Each ACLED country
        # workbook has ~10–25 years × 12 months = 120–300 unique keys, so
        # the accumulator never exceeds a few KB even though the file
        # itself is megabytes.
        bucket: Dict[str, Dict[str, int]] = {}
        for row in rows_iter:
            if row is None:
                continue
            try:
                year_v  = row[year_i]
                month_v = row[month_i]
                ev_v    = row[ev_i] or 0
                fat_v   = row[fat_i] if (fat_i is not None and fat_i < len(row)) else 0

                if isinstance(month_v, (int, float)):
                    m = int(month_v)
                elif isinstance(month_v, str):
                    m = _HDX_MONTH_LOOKUP.get(month_v.strip().lower())
                    if m is None:
                        continue
                else:
                    continue
                y = int(year_v) if year_v is not None else None
                if y is None:
                    continue

                key = f"{y:04d}-{m:02d}"
                slot = bucket.setdefault(key, {"events": 0, "fatalities": 0})
                slot["events"]     += int(ev_v or 0)
                slot["fatalities"] += int(fat_v or 0)
            except (TypeError, ValueError):
                continue
    finally:
        # close() releases the underlying ZipFile and the row iterator.
        # Important on Render: without this, the workbook handle (and its
        # internal XML parser state) lingers until the next gc cycle.
        try:
            wb.close()
        except Exception:
            pass

    out = [
        {"month": k, "events": v["events"], "fatalities": v["fatalities"], "country": country_label}
        for k, v in bucket.items()
    ]
    out.sort(key=lambda r: r["month"])
    return out


def fetch_hdx_event_counts(force: bool = False) -> Dict[str, Any]:
    """Live monthly ACLED event counts for every chokepoint-relevant country.

    No authentication required — pulls the public ACLED→HDX mirror that
    refreshes every Wednesday. Cache TTL is 6 hours (HDX is stable between
    refreshes and we don't want to thrash their API).

    Returns:
        {
          "fetched_utc": ISO-8601,
          "newest_month": "YYYY-MM",        # max across all countries
          "by_country": {
            "yemen": [{month, events, fatalities, country}, ...],
            "iran":  [...],
            ...
          },
          "errors": {country: error_str, ...}
        }
    """
    cache_key = "hdx_event_counts"
    if not force:
        cached = _read_cache(cache_key, 6 * 3600)
        if cached:
            return cached
    # Singleflight the actual fetch (5 sequential XLSX downloads ~ 30-60s).
    # Concurrent callers share one fetch instead of all hammering HDX.
    sf_key = f"hdx_event_counts:{'force' if force else 'normal'}"
    return _sf.run(sf_key, lambda: _do_fetch_hdx_event_counts(force, cache_key))


def _do_fetch_hdx_event_counts(force: bool, cache_key: str) -> Dict[str, Any]:
    # Re-check cache after singleflight handoff (a leader that finished
    # while we waited may have populated the cache file).
    if not force:
        cached = _read_cache(cache_key, 6 * 3600)
        if cached:
            return cached

    out: Dict[str, Any] = {
        "fetched_utc": datetime.utcnow().isoformat() + "Z",
        "by_country": {},
        "by_country_meta": {},
        "errors": {},
    }
    newest = ""

    # Read the previous snapshot so per-country failures can fall back to
    # last-known-good data rather than silently dropping that country from
    # the aggregate. Without this, a single corrupt Yemen XLSX (HDX returns
    # an HTML error page → "Bad magic number for file header") makes the
    # §01 chart's bars drop ~17% because Yemen's contribution disappears
    # until the next refresh.
    prev_snap = _read_stale_cache(cache_key) or {}
    prev_by_country = prev_snap.get("by_country") or {}
    prev_meta = prev_snap.get("by_country_meta") or {}

    # Sequential fetch — 5 countries × ~2MB XLSX each. Sequential keeps us
    # under HDX's rate limits AND under Render's 512MB memory ceiling
    # (parallel parsing peaks at ~5x the per-country footprint).
    import gc as _gc
    for label, pkg in _HDX_PACKAGES.items():
        meta = _hdx_resource_url(pkg)
        if not meta or not meta.get("url"):
            out["errors"][label] = "package_show returned no political_violence resource"
            # Reuse last known good for this country so the aggregate stays whole
            if label in prev_by_country:
                out["by_country"][label] = prev_by_country[label]
                out["by_country_meta"][label] = {**prev_meta.get(label, {}), "served_from": "stale_cache"}
                cur = prev_by_country[label][-1]["month"] if prev_by_country[label] else ""
                if cur and cur > newest:
                    newest = cur
            continue
        r = None
        content = None
        rows = []
        # One inline retry for the body fetch — HDX intermittently returns
        # truncated or HTML-wrapped responses through their CDN. A single
        # retry after a small backoff usually clears it without doubling
        # the request rate enough to trip rate limits.
        last_err = None
        for attempt in range(2):
            try:
                r = requests.get(meta["url"], headers=_BROWSER_HEADERS,
                                 timeout=30, allow_redirects=True)
                if not r.ok:
                    last_err = f"HTTP {r.status_code}"
                    continue
                # Validate the body is actually an XLSX (PK ZIP magic) before
                # handing to openpyxl. HTML error pages from HDX/Cloudflare
                # would otherwise produce "Bad magic number" in the parser.
                content = r.content
                if not content.startswith(b"PK"):
                    ctype = r.headers.get("Content-Type", "?")
                    last_err = f"non-XLSX response (ct={ctype}, first 8 bytes={content[:8]!r})"
                    if attempt == 0:
                        logger.warning(f"HDX {label}: {last_err}, retrying once")
                        continue
                    break
                rows = _parse_hdx_monthly_xlsx(content, label)
                if not rows:
                    last_err = "parser returned 0 rows"
                    if attempt == 0:
                        continue
                last_err = None
                break
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                if attempt == 0:
                    logger.warning(f"HDX {label}: {last_err}, retrying once")
                    continue
        try:
            if rows:
                out["by_country"][label] = rows
                out["by_country_meta"][label] = {
                    "resource_name": meta.get("name"),
                    "last_modified": meta.get("last_modified"),
                    "row_count": len(rows),
                }
                cur = rows[-1]["month"]
                if cur > newest:
                    newest = cur
            else:
                # Fetch failed after retries → reuse last known good data
                # for this country instead of dropping it from the aggregate.
                out["errors"][label] = last_err or "unknown failure"
                logger.warning(f"HDX {label} fetch failed: {last_err}")
                if label in prev_by_country:
                    out["by_country"][label] = prev_by_country[label]
                    prior = prev_meta.get(label, {})
                    out["by_country_meta"][label] = {**prior, "served_from": "stale_cache", "stale_reason": last_err}
                    cur = prev_by_country[label][-1]["month"] if prev_by_country[label] else ""
                    if cur and cur > newest:
                        newest = cur
        finally:
            # Drop the response body and any intermediate bytes immediately
            # so the next country's parse doesn't compound memory use.
            del r, content
            _gc.collect()

    out["newest_month"] = newest or None
    if out["by_country"]:
        _write_cache(cache_key, out)
    else:
        # Don't cache an all-errors snapshot — it'd block recovery for 6h.
        # But return a stale cache if available so the UI doesn't go blank.
        stale = _read_stale_cache(cache_key)
        if stale:
            stale["served_stale"] = True
            return stale
    return out


# ─── Master Dataset (CSV Backbone) ──────────────────────────────────────────

def _extend_master_timeseries_with_live(result: dict) -> None:
    """Append live daily rows to the static master timeseries (in-place).

    The static `data/master_dataset.json` is built when the developer last
    regenerated it; without intervention it stays frozen at that date even
    when the live Brent / DXY / OVX / SPR caches are weeks ahead. This
    function fills the gap so the Volatility / Time-Series / Scatter charts
    extend all the way to today's print.

    Live sources used (read from on-disk caches — never triggers a fetch):
        Brent_Price          ← brent_prices.json (daily, EIA + CSV fallback)
        DXY                  ← dxy.json (daily, yfinance)
        OVX                  ← ovx.json (daily, yfinance)
        SPR_Release_Volume   ← spr_data.json (weekly carry-forward)
        China_PMI            ← china_pmi.json (monthly carry-forward)

    Daily_Volatility for new rows is computed as the rolling 5-day standard
    deviation of Brent log returns × √252 to match the static file's units.
    Attack-frequency and event-flag columns are set to 0 / NaN for new rows;
    the live HDX monthly counts are surfaced separately via /api/live-event-counts
    (we don't fabricate daily granularity from monthly aggregates).

    Mutates `result["timeseries"]` in place; no-op if no rows would be added.
    """
    ts = result.get("timeseries") or []
    if not ts:
        return
    last_static_date = ts[-1].get("date")
    if not last_static_date:
        return

    # Read live caches without triggering network IO. If any of these are
    # missing we still proceed with whatever IS available; the merge handles
    # missing per-row fields gracefully.
    brent = _read_cache("brent_prices",  config.CACHE_TTL_BRENT)    or []
    dxy   = _read_cache("dxy",            config.CACHE_TTL_YFINANCE) or []
    ovx   = _read_cache("ovx",            config.CACHE_TTL_YFINANCE) or []
    spr   = _read_cache("spr_data",       86_400 * 14)               or []
    pmi   = _read_cache("china_pmi",      config.CACHE_TTL_FRED)     or []

    if not brent:
        return  # no live data → nothing to extend with

    # Index everything by date for O(1) lookup.
    def _by_date(rows, val_key):
        return {r["date"]: r.get(val_key) for r in rows if r.get("date")}

    brent_by = _by_date(brent, "price")
    dxy_by   = _by_date(dxy,   "value")
    ovx_by   = _by_date(ovx,   "value")
    spr_by   = _by_date(spr,   "value") if spr else {}
    # China PMI is monthly (YYYY-MM-01); we use last-known carry-forward.
    pmi_sorted = sorted([(r["date"][:10], r.get("value")) for r in pmi if r.get("date")])

    def _pmi_for(date):
        # Most recent PMI observation on/before `date`.
        last = None
        for d, v in pmi_sorted:
            if d <= date:
                last = v
            else:
                break
        return last

    # SPR is weekly with carry-forward semantics.
    spr_sorted = sorted([(d, v) for d, v in spr_by.items()])

    def _spr_for(date):
        last = 0
        for d, v in spr_sorted:
            if d <= date:
                last = v
            else:
                break
        return last

    # Build the union of dates we have live data for, beyond last_static.
    new_dates = sorted({d for d in brent_by if d > last_static_date})
    if not new_dates:
        return

    # Daily_Volatility: rolling 5-day std of log returns. Seed with the last
    # 5 prices from the static series to keep the boundary smooth.
    import math
    seed = []
    for r in ts[-6:]:
        if r.get("brent_price") is not None:
            seed.append(r["brent_price"])
    rolling = list(seed)

    def _vol_after_appending(price):
        rolling.append(price)
        if len(rolling) < 6:
            return None
        window = rolling[-6:]
        rets = []
        for i in range(1, len(window)):
            if window[i-1] and window[i] and window[i-1] > 0:
                rets.append(math.log(window[i] / window[i-1]))
        if len(rets) < 2:
            return None
        mean = sum(rets) / len(rets)
        var = sum((x - mean) ** 2 for x in rets) / (len(rets) - 1)
        return round((var ** 0.5) * (252 ** 0.5), 4)

    # ── Live event counts from HDX (monthly aggregates) ────────────────
    # The thesis CSV stops publishing weekly_attacks/fatalities_count past
    # its observation window (Sep 2025). Without filling these, the main
    # "Price vs. Conflict Intensity" chart shows live Brent through April
    # 2026 but ZERO bars across the entire US-Iran war period — visually
    # implying the war is silent, which is the opposite of the truth.
    #
    # The dashboard tracks BOTH chokepoints — Bab (Houthi/Yemen theatre)
    # and Hormuz (Iran/IRGC theatre). The live signal therefore combines
    # Yemen + Iran HDX monthly conflict events. This is intentionally a
    # broader scope than the thesis's curated "Houthi maritime strikes"
    # subset (which peaked at 47/week) because the conflict itself has
    # broadened: a chart that only counted Houthi maritime ops during a
    # multi-front US-Iran war would be misleading by omission.
    #
    # The boundary is real and the chart subtitle discloses it: pre-Oct
    # 2025 = thesis WeeklyAttackFreq (Houthi maritime); post-Oct 2025 =
    # HDX Yemen + Iran combined. The y-axis scale will jump because the
    # war IS bigger — that's the honest representation.
    hdx_snap = _read_cache("hdx_event_counts", 86_400 * 7) or {}
    hdx_by = hdx_snap.get("by_country") or {}

    # Build {YYYY-MM: {events_per_day, fatalities_per_day}} for both
    # chokepoint theatres combined. Per-chokepoint live cards still
    # surface the country-specific counts separately.
    monthly_per_day: Dict[str, Dict[str, float]] = {}
    for country_label in ("yemen", "iran"):
        for r in (hdx_by.get(country_label) or []):
            ym = r.get("month")
            if not ym:
                continue
            try:
                y, m = int(ym[:4]), int(ym[5:7])
                # Days in month — handles leap years
                if m == 12:
                    next_month = datetime(y + 1, 1, 1)
                else:
                    next_month = datetime(y, m + 1, 1)
                days_in_month = (next_month - datetime(y, m, 1)).days
            except Exception:
                continue
            slot = monthly_per_day.setdefault(ym, {"events": 0.0, "fatalities": 0.0, "days": days_in_month})
            slot["events"]     += float(r.get("events", 0) or 0)
            slot["fatalities"] += float(r.get("fatalities", 0) or 0)

    def _hdx_for_day(date_str):
        ym = date_str[:7]
        slot = monthly_per_day.get(ym)
        if not slot or not slot["days"]:
            return None, None
        # Daily share = monthly total / days_in_month
        return (slot["events"] / slot["days"], slot["fatalities"] / slot["days"])

    # Schema-match the static rows so downstream code doesn't have to handle
    # missing keys. Use the last static row as the field template.
    template_keys = list(ts[-1].keys())

    # War onset for the IranIsrael_Escalation flag — set to 1 from this
    # date forward in live-extension rows so charts that gate on the war
    # phase don't report "war = false" during the war.
    WAR_ONSET = "2026-02-28"

    appended = 0
    weekly_buf: List[float] = []  # rolling 7-day attack window
    for d in new_dates:
        bp = brent_by[d]
        if bp is None:
            continue
        row = {k: None for k in template_keys}
        row["date"] = d
        row["brent_price"] = round(float(bp), 2)
        row["daily_volatility"] = _vol_after_appending(float(bp))
        row["dxy"]               = round(float(dxy_by[d]), 2) if dxy_by.get(d) is not None else None
        row["ovx"]               = round(float(ovx_by[d]), 2) if ovx_by.get(d) is not None else None
        row["spr_release_volume"] = _spr_for(d)
        row["china_pmi"]         = _pmi_for(d)

        # Daily attack count — distribute the HDX monthly total across that
        # month's days. Non-integer estimate; we round only at the weekly
        # rollup so accumulated fractions don't all snap to zero.
        daily_ev, daily_fat = _hdx_for_day(d)
        if daily_ev is not None:
            weekly_buf.append(daily_ev)
            if len(weekly_buf) > 7:
                weekly_buf.pop(0)
            row["daily_attacks"] = int(round(daily_ev))
            row["weekly_attacks"] = int(round(sum(weekly_buf)))
            row["fatalities_count"] = int(round(daily_fat or 0))
            # We don't know the per-event tanker / chokepoint flags from
            # HDX (only country-level counts), so leave those at 0 rather
            # than fabricate a split.
            row["tanker_attacks"] = 0
            row["chokepoint_attacks"] = 0
        else:
            row["weekly_attacks"]    = 0
            row["daily_attacks"]     = 0
            row["fatalities_count"]  = 0
            row["tanker_attacks"]    = 0
            row["chokepoint_attacks"] = 0

        # Other 0/1 indicator columns
        row["opec_dummy"]            = 0
        row["russia_ukraine_dummy"]  = 0
        row["opec_decision"]         = 0
        row["russia_ukraine_attacks"] = 0
        row["iran_israel_escalation"] = 1 if d >= WAR_ONSET else 0
        ts.append(row)
        appended += 1

    if appended:
        logger.info(f"Master timeseries: extended with {appended} live rows ({new_dates[0]} → {new_dates[-1]})")
        # Update KPI block so /api/master is internally consistent.
        result.setdefault("kpis", {})
        result["kpis"]["timeseries_newest_date"] = new_dates[-1]
        result["kpis"]["timeseries_extended_rows"] = appended


_master_memo: Optional[dict] = None
_master_memo_ts: float = 0.0
_MASTER_MEMO_TTL = 90  # seconds — short so live-extension picks up new Brent prints quickly


def load_master_dataset() -> dict:
    """Load the master dataset CSV and return as structured JSON.

    In-process memo (90s TTL) so /api/master, /api/comparative, and the
    repeated frontend polls don't each re-parse the 190KB static JSON +
    re-run the live extender. Without the memo each call materialised a
    fresh ~5MB Python dict tree from disk; under concurrent traffic on
    Render's 512MB free tier this contributed to OOMs.

    Fast path: if data/master_dataset.json (pre-computed at build / committed
    to repo) exists, serve it directly. This avoids cold-start pandas work on
    Render's free tier where the live computation can exceed Cloudflare's
    edge timeout, leaving the cache permanently un-warmed. The static file
    only needs to be regenerated when the underlying CSV changes — live
    Brent / DXY / OVX KPIs are still pulled from their dedicated cached
    endpoints by the frontend.
    """
    global _master_memo, _master_memo_ts
    if _master_memo is not None and (time.time() - _master_memo_ts) < _MASTER_MEMO_TTL:
        return _master_memo

    static_json = config.DATA_DIR / "master_dataset.json"
    if static_json.exists():
        try:
            with open(static_json, "r") as f:
                result = json.load(f)
            # Refresh the live KPI fields from caches if available so the
            # static file doesn't go stale on price snapshots.
            try:
                live_brent = _read_cache("brent_prices", config.CACHE_TTL_BRENT)
                if live_brent and len(live_brent) >= 2:
                    result.setdefault("kpis", {})
                    result["kpis"]["latest_brent_price"] = round(live_brent[-1]["price"], 2)
                    result["kpis"]["brent_price_change"] = round(live_brent[-1]["price"] - live_brent[-2]["price"], 2)
                live_dxy = _read_cache("dxy", config.CACHE_TTL_YFINANCE)
                if live_dxy:
                    result.setdefault("kpis", {})["latest_dxy"] = round(live_dxy[-1]["value"], 2)
                live_ovx = _read_cache("ovx", config.CACHE_TTL_YFINANCE)
                if live_ovx:
                    result.setdefault("kpis", {})["latest_ovx"] = round(live_ovx[-1]["value"], 2)
            except Exception as e:
                logger.warning(f"master_dataset.json: could not refresh live KPIs: {e}")
            # Extend the timeseries with live daily prices so the main charts
            # stop ending at the static file's frozen date. Without this,
            # the volatility / Brent / DXY / OVX charts dead-end at whatever
            # date the dev last regenerated master_dataset.json — currently
            # 2025-10-01 — even though Brent + macro caches are within days
            # of wall-clock. See _extend_master_timeseries_with_live for
            # the merge logic.
            try:
                _extend_master_timeseries_with_live(result)
            except Exception as e:
                logger.warning(f"master_dataset.json: live timeseries extend failed: {e}")
            _master_memo = result
            _master_memo_ts = time.time()
            return result
        except Exception as e:
            logger.error(f"master_dataset.json read failed, falling back to live: {e}")

    cached = _read_cache("master_dataset", 1800)  # 30 min cache
    if cached:
        return cached

    if not config.MYLES_DATASET_PATH.exists():
        logger.error(f"Master dataset not found: {config.MYLES_DATASET_PATH}")
        return {"timeseries": [], "kpis": {}, "price_windows": {}, "correlation": []}

    df = pd.read_csv(config.MYLES_DATASET_PATH).copy()
    date_col = df.columns[0]
    df = df.rename(columns={date_col: "Date"}).copy()
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").copy()

    # Time series records — vectorized (avoids slow iterrows)
    col_map = {
        "Brent_Price": ("brent_price", 2, False),
        "Daily_Volatility": ("daily_volatility", 4, False),
        "DXY": ("dxy", 2, False),
        "OVX": ("ovx", 2, False),
        "SPR_Release_Volume": ("spr_release_volume", 4, False),
        "China_PMI": ("china_pmi", 2, True),        # True = also exclude zeros
        "Baker_Hughes_Rigs": ("baker_hughes_rigs", 1, True),
    }
    int_cols = {
        "WeeklyAttackFreq": "weekly_attacks",
        "daily_attacks": "daily_attacks",
        "tanker_attacks": "tanker_attacks",
        "chokepoint_attacks": "chokepoint_attacks",
        "fatalities": "fatalities_count",
        "OPEC_Dummy": "opec_dummy",
        "RussiaUkraine_Dummy": "russia_ukraine_dummy",
        "OPEC_Decision": "opec_decision",
        "RussiaUkraine_Attacks": "russia_ukraine_attacks",
        "IranIsrael_Escalation": "iran_israel_escalation",
    }
    ts = pd.DataFrame({"date": df["Date"].dt.strftime("%Y-%m-%d")})
    for src, (dst, rnd, excl_zero) in col_map.items():
        if src in df.columns:
            vals = df[src].round(rnd)
            if excl_zero:
                vals = vals.where(df[src] != 0)
            ts[dst] = vals
        else:
            ts[dst] = np.nan
    for src, dst in int_cols.items():
        ts[dst] = df[src].fillna(0).astype(int) if src in df.columns else 0

    # Convert to list of dicts, replacing NaN/numpy types with JSON-safe Python types
    def _clean(v):
        if v is None:
            return None
        if isinstance(v, (np.floating, float)):
            return None if np.isnan(v) else round(float(v), 6)
        if isinstance(v, (np.integer, int)):
            return int(v)
        return v
    timeseries = [{k: _clean(v) for k, v in row.items()} for row in ts.to_dict("records")]

    # KPIs
    valid_prices = df["Brent_Price"].dropna()

    # Use cached Brent price if available (non-blocking); fall back to CSV last-row
    live_brent = _read_cache("brent_prices", config.CACHE_TTL_BRENT)
    if live_brent and len(live_brent) >= 2:
        live_latest = live_brent[-1]["price"]
        live_prev = live_brent[-2]["price"]
        latest_brent = round(live_latest, 2)
        brent_change = round(live_latest - live_prev, 2)
    else:
        latest_brent = round(float(valid_prices.iloc[-1]), 2)
        brent_change = round(float(valid_prices.iloc[-1] - valid_prices.iloc[-2]), 2) if len(valid_prices) > 1 else 0

    # Use cached DXY (non-blocking); fall back to CSV last-row
    live_dxy = _read_cache("dxy", config.CACHE_TTL_YFINANCE)
    latest_dxy = round(live_dxy[-1]["value"], 2) if live_dxy else (
        round(float(df["DXY"].dropna().iloc[-1]), 2) if df["DXY"].dropna().shape[0] > 0 else None
    )

    # Use cached OVX (non-blocking); fall back to CSV last-row
    live_ovx = _read_cache("ovx", config.CACHE_TTL_YFINANCE)
    latest_ovx = round(live_ovx[-1]["value"], 2) if live_ovx else (
        round(float(df["OVX"].dropna().iloc[-1]), 2) if df["OVX"].dropna().shape[0] > 0 else None
    )

    kpis = {
        "avg_brent_price": round(float(valid_prices.mean()), 2),
        "latest_brent_price": latest_brent,
        "brent_price_change": brent_change,
        "peak_volatility": round(float(df["Daily_Volatility"].max()), 4),
        "max_weekly_attacks": int(df["WeeklyAttackFreq"].max()),
        "latest_dxy": latest_dxy,
        "latest_ovx": latest_ovx,
        "total_trading_days": 505,  # Regression observations (506 price days minus 1 for return calc)
    }

    # Price windows (event study: T-2 to T+5)
    price_cols = ["Price_T-2", "Price_T-1", "Price_T0", "Price_T+1", "Price_T+2", "Price_T+3", "Price_T+4", "Price_T+5"]
    attack_rows = df[df["WeeklyAttackFreq"] > 0]
    price_windows = {}
    for col in price_cols:
        if col in df.columns:
            values = attack_rows[col].replace(0, np.nan).dropna()
            price_windows[col] = round(float(values.mean()), 2) if len(values) > 0 else 0

    # Correlation matrix.
    # The OPEC / Russia-Ukraine / Iran-Israel dummy variables are zero across
    # the entire sample window (the events they encode either pre-date or
    # post-date the regression sample), which makes their pairwise correlations
    # NaN → rendered as a column of zeros in the heatmap. That's an
    # embarrassing visual ("everything correlates 0.00 with X?") so we drop
    # any column with zero variance before building the matrix. Logged so it's
    # visible during refreshes if a dummy ever does activate.
    corr_cols = ["Brent_Price", "Daily_Volatility", "WeeklyAttackFreq", "DXY", "OVX",
                 "OPEC_Dummy", "RussiaUkraine_Dummy", "IranIsrael_Escalation",
                 "China_PMI", "Baker_Hughes_Rigs", "SPR_Release_Volume"]
    available_cols = [c for c in corr_cols if c in df.columns]
    # Drop columns where every value is the same (variance == 0). The
    # `replace(0, NaN).dropna(how="all")` row-filter below would already drop
    # all-zero rows, but a column where ALL non-zero values are identical (or
    # the column is entirely NaN/0) still produces a column-of-zeros after
    # corr(). Filtering at the column level is the correct fix.
    nonconstant_cols = [c for c in available_cols if df[c].nunique(dropna=True) > 1]
    dropped = [c for c in available_cols if c not in nonconstant_cols]
    if dropped:
        logger.info(f"Correlation matrix: dropped zero-variance columns {dropped}")
    corr_df = df[nonconstant_cols].replace(0, np.nan).dropna(how="all").corr()
    correlation = {
        "labels": list(corr_df.columns),
        "matrix": [[round(float(v), 3) if pd.notna(v) else 0 for v in row] for row in corr_df.values],
        "dropped_constant": dropped,
    }

    result = {
        "timeseries": timeseries,
        "kpis": kpis,
        "price_windows": price_windows,
        "correlation": correlation,
    }
    _write_cache("master_dataset", result)
    return result


# ─── Hypothesis Results (hardcoded from notebook) ───────────────────────────

# ─── Iran Events (Current Events Tab) ────────────────────────────────────

def fetch_iran_events() -> List[dict]:
    """Fetch Iran-related events from ACLED API with cache.

    Singleflight-wrapped: concurrent callers share one in-flight fetch
    instead of all racing past the cache check and making parallel ACLED
    queries. Without this, /api/iran-events under load was simultaneously
    fetching live (1173 events, slow) AND serving the 740-event JSON
    fallback from another caller — which led to inconsistent UI state.
    """
    cached = _read_cache("iran_events", 3600)  # 1-hour cache
    if cached:
        logger.info("Iran events: serving from cache")
        return cached
    return _sf.run("iran_events", _do_fetch_iran_events)


def _do_fetch_iran_events() -> List[dict]:
    global _iran_fetch_error
    # Re-check the cache after the singleflight handoff so a leader that
    # finished while we were waiting doesn't make us re-fetch.
    cached = _read_cache("iran_events", 3600)
    if cached:
        _iran_fetch_error = None
        return cached

    try:
        token = _get_acled_token()
        iran_fields = "event_id_cnty|event_date|event_type|sub_event_type|actor1|actor2|location|latitude|longitude|notes|fatalities|tags"

        # 60s + one retry to match _paginated_acled_fetch — ACLED's first page
        # request from a cold connection (Render) frequently times out at 30s
        # and silently puts us on the 2025-03 fallback. The retry keeps live
        # data flowing without a manual /api/refresh.
        def _acled_get_with_retry(params):
            last_err = None
            for attempt in range(2):
                try:
                    r = requests.get(
                        config.ACLED_DATA_URL,
                        headers={**_BROWSER_HEADERS, "Authorization": f"Bearer {token}"},
                        params=params,
                        timeout=60,
                    )
                    r.raise_for_status()
                    return r
                except (requests.Timeout, requests.ConnectionError) as e:
                    last_err = e
                    if attempt == 0:
                        logger.warning(f"Iran ACLED: {type(e).__name__}, retrying once")
            raise last_err

        def _fetch_iran_country():
            results = []
            for page in range(1, 8):  # Cap at 7 pages (35k events max — more than enough)
                resp = _acled_get_with_retry({
                    "_format": "json", "country": "Iran",
                    "event_date": "2025-01-01|2026-12-31", "event_date_where": "BETWEEN",
                    "fields": iran_fields, "limit": 5000, "page": page,
                })
                batch = resp.json().get("data", [])
                if not batch:
                    break
                results.extend(batch)
                if len(batch) < 5000:
                    break
            logger.info(f"Iran events (country): {len(results)} events")
            return results

        def _fetch_iran_bilateral(actor1, actor2):
            resp = _acled_get_with_retry({
                "_format": "json", "actor1": actor1, "actor1_where": "LIKE",
                "actor2": actor2, "actor2_where": "LIKE",
                "event_date": "2025-01-01|2026-12-31", "event_date_where": "BETWEEN",
                "fields": iran_fields, "limit": 5000,
            })
            bilateral = resp.json().get("data", [])
            logger.info(f"Iran bilateral ({actor1}→{actor2}): {len(bilateral)} events")
            return bilateral

        # Run all 3 queries in parallel
        with ThreadPoolExecutor(max_workers=3) as pool:
            f_country = pool.submit(_fetch_iran_country)
            f_us_iran = pool.submit(_fetch_iran_bilateral, "United States", "Iran")
            f_iran_us = pool.submit(_fetch_iran_bilateral, "Iran", "United States")

        all_events = []
        seen_ids = set()
        for batch in [f_country.result(), f_us_iran.result(), f_iran_us.result()]:
            for e in batch:
                eid = e.get("event_id_cnty")
                if eid and eid not in seen_ids:
                    all_events.append(e)
                    seen_ids.add(eid)

        _iran_fetch_error = None
        if all_events:
            _write_cache("iran_events", all_events)
            logger.info(f"Iran events: total {len(all_events)} events fetched and cached")
            return all_events
        else:
            _iran_fetch_error = "ACLED returned 0 events for all Iran queries"
            logger.warning(_iran_fetch_error)

    except Exception as e:
        _iran_fetch_error = f"{type(e).__name__}: {e}"
        logger.warning(f"Iran events API failed: {_iran_fetch_error}")

    # Fallback: load from pre-fetched JSON file
    fallback = _load_iran_json_fallback()
    if fallback:
        _iran_fetch_error = None  # Clear error since fallback succeeded
    return fallback


_iran_fallback_memo: Optional[List[dict]] = None

def _load_iran_json_fallback() -> List[dict]:
    """Load Iran events from local JSON fallback file.

    Memoized after first successful parse — the 400KB file materializes to
    ~2MB of dicts and was previously being reparsed on every /api/iran-events
    and /api/iran-impact call.
    """
    global _iran_fallback_memo
    if _iran_fallback_memo is not None:
        return _iran_fallback_memo
    path = config.DATA_DIR / "iran_events.json"
    if not path.exists():
        logger.warning("Iran events: no JSON fallback file found")
        return []
    try:
        events = json.loads(path.read_text())
        _iran_fallback_memo = events
        logger.info(f"Iran events: loaded {len(events)} events from JSON fallback")
        return events
    except Exception as e:
        logger.warning(f"Iran events JSON fallback failed: {e}")
        return []


def get_iran_fetch_error() -> Optional[str]:
    return _iran_fetch_error


def get_curated_iran_events() -> List[dict]:
    """Return curated timeline of major US-Iran events (2025-2026) with coordinates.

    Severity scale (1-5):
        5 = Civilization-altering (nuclear/WMD use, total global energy shutdown)
        4 = Major war escalation (war begins, Hormuz closed, massive retaliation, 100+ killed/day)
        3 = Significant military action (major strikes, infrastructure hit, new fronts open)
        2 = Provocations & buildups (tanker seizures, force deployments, sanctions)
        1 = Diplomatic/rhetorical (talks, threats, policy shifts)
    """
    return [
        # ── Phase 1: Maximum Pressure Restored (Jan-Feb 2025) ──
        {"date": "2025-01-20", "title": "Trump Inaugurated, Rescinds Biden-Era Iran Policies", "type": "diplomatic", "description": "Trump signs EO 14148 rescinding 67 Biden-era executive orders including Iran sanctions-related actions.", "severity": 1, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},
        {"date": "2025-02-04", "title": "Trump Signs 'Maximum Pressure' Executive Order", "type": "sanctions", "description": "NSPM-2 restores maximum pressure campaign: Treasury to impose maximum economic pressure, State Dept rescinds sanctions waivers, campaign to drive Iran oil exports to zero.", "severity": 2, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},
        {"date": "2025-02-06", "title": "OFAC Sanctions Iranian Oil Shipping Network", "type": "sanctions", "description": "Treasury's OFAC sanctions international network of parties and vessels facilitating Iranian crude oil shipments to China.", "severity": 1, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},

        # ── Phase 2: Nuclear Talks Begin (Mar-Jun 2025) ──
        {"date": "2025-03-07", "title": "Trump Sends Letter to Khamenei with 2-Month Deadline", "type": "diplomatic", "description": "Trump sends letter via UAE diplomat Anwar Gargash proposing nuclear negotiations, warning of military consequences if rejected.", "severity": 1, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2025-04-12", "title": "Round 1: US-Iran Indirect Talks Begin in Muscat", "type": "diplomatic", "description": "First indirect US-Iran nuclear talks mediated by Oman. US envoy Witkoff and Iranian FM Araghchi in separate rooms. Both sides call talks 'constructive.'", "severity": 1, "lat": 23.5880, "lon": 58.3829, "location": "Muscat, Oman"},
        {"date": "2025-05-11", "title": "Round 4: US Demands Complete Dismantlement", "type": "diplomatic", "description": "Fourth round in Muscat. Witkoff demands complete dismantlement of Natanz, Fordow, and Isfahan. Positions harden.", "severity": 1, "lat": 23.5880, "lon": 58.3829, "location": "Muscat, Oman"},
        {"date": "2025-05-31", "title": "IAEA: Iran Has 400+ kg of 60% Enriched Uranium", "type": "nuclear", "description": "Confidential IAEA report confirms 400+ kg of 60% enriched uranium, enough for ~10 nuclear weapons if further enriched; total stockpile 40x JCPOA limit.", "severity": 3, "lat": 33.5103, "lon": 51.9250, "location": "Natanz, Iran"},

        # ── Phase 3: The Twelve-Day War (Jun 2025) ──
        {"date": "2025-06-13", "title": "Israel Launches 'Operation Rising Lion': Strikes Iran", "type": "military", "description": "Israel launches surprise strikes on Iranian nuclear facilities including Natanz. Prominent military leaders and nuclear scientists assassinated. US-Iran talks suspended.", "severity": 4, "lat": 33.5103, "lon": 51.9250, "location": "Natanz, Iran"},
        {"date": "2025-06-21", "title": "US Launches 'Operation Midnight Hammer'", "type": "military", "description": "125+ aircraft including seven B-2 bombers with GBU-57 bunker busters strike Fordow, Natanz, Isfahan. Tomahawks from submarines. Trump claims facilities 'obliterated.'", "severity": 4, "lat": 34.7564, "lon": 51.0596, "location": "Fordow, Iran"},
        {"date": "2025-06-22", "title": "Iran Retaliates: 550+ Missiles, 1000+ Drones at Israel", "type": "military", "description": "Iran launches over 550 ballistic missiles and 1,000+ drones at Israeli and US targets. Most intercepted by Israel and US.", "severity": 4, "lat": 32.0853, "lon": 34.7818, "location": "Tel Aviv, Israel"},
        {"date": "2025-06-24", "title": "Twelve-Day War Ceasefire Agreed", "type": "diplomatic", "description": "Israel and Iran agree to ceasefire under US pressure, ending the Twelve-Day War.", "severity": 1, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},

        # ── Phase 4: Snapback Sanctions & Isolation (Aug-Oct 2025) ──
        {"date": "2025-08-28", "title": "E3 Triggers JCPOA Snapback Sanctions Mechanism", "type": "sanctions", "description": "UK, France, and Germany invoke JCPOA snapback citing Iran's 'significant non-performance.' 30-day countdown begins.", "severity": 2, "lat": 40.7489, "lon": -73.9680, "location": "New York, NY (UN)"},
        {"date": "2025-09-27", "title": "UN Snapback Sanctions Formally Reimposed on Iran", "type": "sanctions", "description": "All UN sanctions lifted under JCPOA formally reimposed: travel bans, asset freezes, arms embargo, ballistic missile restrictions. EU follows Sept 29.", "severity": 2, "lat": 40.7489, "lon": -73.9680, "location": "New York, NY (UN)"},
        {"date": "2025-10-18", "title": "Iran Officially Terminates the JCPOA", "type": "diplomatic", "description": "Iran declares the JCPOA over on 'Termination Day.' Iran, Russia, and China declare UN sanctions invalid.", "severity": 2, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},

        # ── Phase 5: Protests & Crackdown (Dec 2025 - Jan 2026) ──
        {"date": "2025-12-28", "title": "Massive Anti-Regime Protests Erupt Across Iran", "type": "proxy", "description": "Protests erupt after rial collapses to 1.4M/$1. Tehran Grand Bazaar strikes spread nationwide. 72% food inflation, post-war devastation, snapback sanctions.", "severity": 2, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-01-08", "title": "Iran's Deadliest Crackdown Since 1979", "type": "military", "description": "Security forces launch massive crackdown. Internet fully cut. Firearms and shotguns with metal pellets used against protesters. Thousands reported killed.", "severity": 3, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-01-23", "title": "Trump Announces Naval 'Armada' Heading to Middle East", "type": "military", "description": "Trump announces USS Abraham Lincoln carrier strike group deployment. F/A-18E Super Hornets, F-35C Lightning IIs, guided-missile destroyers.", "severity": 2, "lat": 25.2854, "lon": 55.3500, "location": "Persian Gulf"},

        # ── Phase 6: War Buildup & Hormuz Provocations (Feb 2026) ──
        {"date": "2026-01-30", "title": "IRGC Seizes South Korean Tanker in Strait of Hormuz", "type": "military", "description": "IRGC Navy commandos fast-rope onto South Korean chemical tanker 'Hankuk Chemi II' in Strait of Hormuz, citing 'environmental violations.' Crew of 20 detained at Bandar Abbas. Seoul condemns 'act of piracy.' Mirrors 2021 tanker seizure. Oil markets spike 3.2%.", "severity": 2, "fatalities": 0, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-02-03", "title": "IRGC Attempts to Board US Tanker; Drone Shot Down", "type": "military", "description": "IRGC Navy attempts to intercept US-flagged tanker in Strait of Hormuz. USS McFaul escorts it to safety. F-35C shoots down Iranian Shahed-136 drone.", "severity": 2, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-02-10", "title": "IRGC Fires Warning Shots at Norwegian Tanker Near Hormuz", "type": "military", "description": "IRGC Navy patrol boat fires warning shots across bow of Norwegian-flagged tanker 'Nordic Spirit' in international waters near Hormuz. USS Carney intervenes; IRGC boats withdraw. Norway summons Iranian ambassador. Third Hormuz shipping incident in 11 days.", "severity": 2, "fatalities": 0, "lat": 26.4800, "lon": 56.3000, "location": "Strait of Hormuz"},
        {"date": "2026-02-06", "title": "Round 6: US-Iran Talks Resume in Muscat", "type": "diplomatic", "description": "First talks since June 2025. US delegation: Witkoff, Kushner, CENTCOM commander Adm. Cooper. Iranian FM Araghchi leads. 'Good start.'", "severity": 1, "lat": 23.5880, "lon": 58.3829, "location": "Muscat, Oman"},
        {"date": "2026-02-13", "title": "USS Gerald R. Ford Redeployed; Trump Signals Regime Change", "type": "military", "description": "Ford redirected to Middle East, largest US force posture since 2003 Iraq War. Trump says regime change would be 'best thing that could happen.'", "severity": 2, "lat": 25.2854, "lon": 55.3500, "location": "Persian Gulf"},
        {"date": "2026-02-14", "title": "Pentagon Prepares 'Weeks-Long Sustained Operations'", "type": "military", "description": "US officials confirm military is preparing for sustained operations against Iran lasting weeks.", "severity": 2, "lat": 38.8719, "lon": -77.0563, "location": "Pentagon, VA"},
        {"date": "2026-02-19", "title": "Trump Gives Iran 10-Day Ultimatum", "type": "diplomatic", "description": "Trump tells Iran to reach a 'meaningful' deal within 10-15 days or 'really bad things' will happen. IRGC conducts live-fire Strait of Hormuz drill.", "severity": 2, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-02-24", "title": "F-22s Deployed to Israel; State of the Union Warning", "type": "military", "description": "12 F-22s deployed to Ovda Airbase, the first US offensive weapons in Israel. Trump vows in SOTU that Iran will never have nuclear weapons.", "severity": 2, "lat": 29.9402, "lon": 34.9358, "location": "Ovda Airbase, Israel"},
        {"date": "2026-02-26", "title": "Final Nuclear Talks Fail: No Deal Reached", "type": "diplomatic", "description": "Round 8 in Geneva. US demands: destroy all enrichment sites, surrender uranium, permanent deal, end proxies. Iran refuses missile restrictions. No agreement.", "severity": 2, "lat": 46.2044, "lon": 6.1432, "location": "Geneva, Switzerland"},
        {"date": "2026-02-27", "title": "IAEA Reveals Hidden Uranium; Embassies Evacuate Iran", "type": "nuclear", "description": "IAEA reports 440.9 kg of 60% enriched uranium hidden in Isfahan tunnels. Embassies evacuate. Trump gives go order for Operation Epic Fury from Air Force One.", "severity": 3, "lat": 32.6546, "lon": 51.6680, "location": "Isfahan, Iran"},

        # ── Phase 7: Operation Epic Fury / Iran War (Feb 28+, 2026) ──
        {"date": "2026-02-28", "title": "Operation Epic Fury Begins: US & Israel Strike Iran; Khamenei Killed", "type": "military", "description": "Joint US-Israeli strikes at 2:30 AM EST. ~900 US strikes in 12 hours, 1,000+ targets in 24h. Supreme Leader Khamenei killed in Israeli strikes on Tehran compound. Iran retaliates with dozens of ballistic missiles and drones at Israel and US bases. US Embassy in Kuwait hit.", "severity": 4, "fatalities": 354, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-03-01", "title": "Maersk Suspends Strait of Hormuz Transit; 4 US Soldiers Killed", "type": "proxy", "description": "Maersk suspends all Strait of Hormuz crossings, reroutes around Cape of Good Hope. Tanker transits drop from 24/day to 4. Four US soldiers killed in Kuwait drone strike.", "severity": 3, "fatalities": 4, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-03-02", "title": "IRGC Closes Strait of Hormuz; Hezbollah Enters War", "type": "military", "description": "IRGC officially closes Strait of Hormuz, threatens any ship that passes. 150+ ships anchored outside. Hezbollah fires rockets at Israel; IDF invades southern Lebanon.", "severity": 4, "fatalities": 85, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-03-03", "title": "Global Shipping Suspended; Brent Surges 7.8%", "type": "proxy", "description": "CMA CGM, Hapag-Lloyd, MSC suspend strait transits. IRGC deploys anti-ship missile launchers to Qeshm Island at narrowest Hormuz point. Iranian drones hit Gulf infrastructure. Brent settles at $83.28 (+7.8%).", "severity": 3, "fatalities": 45, "lat": 26.8500, "lon": 55.9000, "location": "Strait of Hormuz"},
        {"date": "2026-03-04", "title": "IRGC Fast Boats Attack Greek Tanker in Hormuz", "type": "military", "description": "IRGC Navy fires RPGs at Greek-flagged VLCC 'Athena Glory' transiting Hormuz under US escort. USS Mason sinks two IRGC boats. First direct naval engagement since war began. Iranian civilian toll passes 1,100. Brent settles at $81.56 (-2.1%) as markets interpret US naval dominance as stabilizing.", "severity": 3, "fatalities": 128, "lat": 26.4500, "lon": 56.4000, "location": "Strait of Hormuz"},
        {"date": "2026-03-05", "title": "IRGC Mines Sink Neutral Vessel; Insurance Withdrawn for Hormuz", "type": "military", "description": "Indian-flagged MV Ganges Spirit strikes Iranian mine in Hormuz, first neutral vessel sunk. Lloyd's suspends all Hormuz hull coverage. Iran fires 500+ missiles, 2,000 drones total. NATO intercepts missile over Turkey. Brent at $88.59 (+8.6%, ~24% since war began).", "severity": 4, "fatalities": 154, "lat": 26.5200, "lon": 56.3500, "location": "Strait of Hormuz"},
        {"date": "2026-03-06", "title": "Iran Strikes Gulf States; Brent Surges to $95.74 (+8.1%)", "type": "military", "description": "Iran strikes Saudi Arabia, Kuwait, Qatar, Bahrain, UAE. Missile hits Jerusalem. US Navy SEALs board IRGC minelayer deploying mines in Hormuz shipping lanes; 23 mines neutralized. CENTCOM declares limited maritime corridor. Brent settles at $95.74 (+8.1%).", "severity": 4, "fatalities": 95, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-03-07", "title": "Iranian Submarine Fires Torpedo at US Destroyer", "type": "military", "description": "IRIN Kilo-class submarine fires torpedo at USS Halsey in Gulf of Oman — first submarine attack on US Navy since WWII. Torpedo malfunctions; P-8 Poseidon forces submarine to surface. Iran's missile capability reported down 90% by Pentagon.", "severity": 3, "fatalities": 180, "lat": 25.5000, "lon": 57.5000, "location": "Gulf of Oman"},
        {"date": "2026-03-08", "title": "Brent Breaks $100/bbl; Shahran Oil Depot and Bandar Abbas Hit", "type": "military", "description": "Israel hits Shahran oil depot near Tehran; toxic smoke over capital. US destroys Bandar Abbas naval base — 14 IRGC vessels sunk. CENTCOM: 'IRGC Navy capability eliminated.' Assembly of Experts names Mojtaba Khamenei new Supreme Leader. Brent breaks $100/bbl.", "severity": 3, "fatalities": 285, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-03-09", "title": "Brent Hits $119/bbl Intraday; IRGC Drones Strike Fujairah Tanker", "type": "military", "description": "IRGC Shahed-136 drones strike Japanese VLCC 'Nippon Maru' at Fujairah, world's 2nd-largest bunkering hub; force majeure declared. Saudi intercepts drone toward Shaybah oilfield. Cumulative toll: 2,400+ killed. Brent hits $119/bbl intraday but settles at $98.96 as Pentagon reports Iran's missile capability down 90% — markets pricing in war's end.", "severity": 3, "fatalities": 252, "lat": 25.1288, "lon": 56.3265, "location": "Fujairah, UAE"},
        {"date": "2026-03-10", "title": "Brent Drops 11.3% as Trump Signals War 'Very Complete'", "type": "diplomatic", "description": "Trump tells CBS war is 'very complete, pretty much.' Energy Secretary's false tanker escort claim triggers flash crash. US destroys 16 Iranian minelayers. Brent drops 11.3% to $87.80, largest single-day drop since March 2022.", "severity": 1, "fatalities": 200, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},
        {"date": "2026-03-11", "title": "Iranian Drones Strike Salalah Port, Oman; Shipping Route Threatened", "type": "military", "description": "Iranian drone strike sets fire to Salalah, Oman's largest commercial port — major escalation hitting neutral mediator state. Threatens critical non-Hormuz shipping route. Iran rejects ceasefire. Pentagon: 140 US troops wounded, 7 killed since Feb 28.", "severity": 3, "fatalities": 180, "lat": 17.0151, "lon": 54.0924, "location": "Salalah, Oman"},
        {"date": "2026-03-12", "title": "Brent Surges Past $105; Hormuz Transit Drops to 2 Ships/Day", "type": "proxy", "description": "Brent hits $105.40/bbl. Hormuz transit at 2 ships/day (was 24). IEA declares first 'severe supply disruption' since 2011. Houthis fire missiles at Saudi Ras Tanura oil terminal. IRGC anti-ship missiles hit tanker near Fujairah. 21% of global oil supply disrupted.", "severity": 4, "fatalities": 70, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-03-13", "title": "Mojtaba Khamenei Threatens Permanent Hormuz Closure", "type": "diplomatic", "description": "New Supreme Leader threatens to permanently seal the Strait of Hormuz. Analysts note IRGC Navy largely destroyed since March 8; uncleared sea mines remain primary residual threat.", "severity": 1, "fatalities": 0, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-03-14", "title": "Trump Orders Strikes on Kharg Island; 90% of Iran's Oil Exports", "type": "military", "description": "Trump orders strikes on Kharg Island, which handles 90% of Iran's crude oil exports. Major escalation targeting Iran's economic lifeline. Separate strikes hit Isfahan, damaging UNESCO-listed Chehel Sotoun Palace.", "severity": 4, "fatalities": 15, "lat": 29.2333, "lon": 50.3167, "location": "Kharg Island, Iran"},
        {"date": "2026-03-15", "title": "Iran FM Declares 'Ready for a Long War'; Rejects Ceasefire", "type": "diplomatic", "description": "Iranian FM Araghchi states Tehran 'never sought a ceasefire' and remains ready for a long war. Trump urges world to keep Hormuz open. Bahrain and Saudi Arabia cancel April F1 Grand Prix over safety.", "severity": 1, "fatalities": 0, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-03-16", "title": "Houthis Await Iran Approval to Resume Red Sea Attacks", "type": "proxy", "description": "Reports emerge that Houthis are awaiting Iranian approval to resume Red Sea shipping attacks if Iran's Hormuz control weakens. FM Araghchi rejects ceasefire. Lebanon death toll reaches 850 with 831,000 displaced.", "severity": 2, "fatalities": 50, "lat": 14.0, "lon": 44.0, "location": "Red Sea"},
        {"date": "2026-03-17", "title": "Israel Assassinates Larijani; Iran Retaliates Across Gulf States", "type": "military", "description": "Israel kills top security official Ali Larijani and Basij commander Gholamreza Soleimani. Iran retaliates with strikes at Saudi Arabia, Kuwait, UAE. Multiple missiles hit Tel Aviv, killing at least 2. Broadening regional escalation.", "severity": 3, "fatalities": 5, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-03-18", "title": "South Pars Gas Field Struck; Intelligence Minister Khatib Killed", "type": "military", "description": "Israel strikes South Pars gas field — world's largest natural gas reserve shared with Qatar. Iran takes several South Pars phases offline. Israel also confirms killing Intelligence Minister Esmail Khatib in Tehran, a day after assassinating Larijani and Basij chief Soleimani. Funeral processions for Larijani draw massive crowds in Tehran. Iran rejects ceasefire; Hormuz transit effectively halted.", "severity": 3, "fatalities": 0, "lat": 27.5000, "lon": 52.0000, "location": "South Pars, Persian Gulf"},

        # ── Day 20 (March 19, 2026) ──
        {"date": "2026-03-19", "title": "Brent Hits $115; Strikes on Yazd Airport; Ras Laffan LNG Hub Hit", "type": "military", "description": "Brent crude surges to $115/bbl after Israeli strikes spark Iranian retaliation on Qatar's Ras Laffan LNG hub. US-Israeli strikes hit Yazd Airport. Trump threatens to 'blow up' South Pars if Iran continues attacking Qatar. Saudi Arabia warns it will take military action against Iran. IDF has carried out 7,600+ strikes across Iran. Pentagon requests $200B+ war budget. FM Araghchi warns allies helping reopen Hormuz risk 'complicity in war crimes.'", "severity": 3, "fatalities": 0, "lat": 31.9049, "lon": 54.2825, "location": "Yazd, Iran"},

        # ── Days 21-30 (March 20-29, 2026) ──
        {"date": "2026-03-20", "title": "Iraq Declares Force Majeure on Basra Crude; Brent at $110.96", "type": "proxy", "description": "Iraq declares force majeure on Basra crude exports citing pipeline damage from Iranian retaliatory strikes near Fao Peninsula. Gulf tanker insurance premiums hit record highs. Brent settles at $110.96 as supply disruptions widen beyond Iran.", "severity": 2, "fatalities": 0, "lat": 30.5085, "lon": 47.7804, "location": "Basra, Iraq"},
        {"date": "2026-03-23", "title": "Brent Drops to $96; IRGC Hormuz 'Tollbooth' Rumors Emerge", "type": "diplomatic", "description": "Brent drops sharply to $96.07 as reports emerge that IRGC may allow select nations to transit Hormuz for fees — a de facto 'tollbooth' system. Markets interpret this as partial de-escalation. China and India reportedly negotiating safe passage for their flagged tankers.", "severity": 1, "fatalities": 0, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-03-25", "title": "Kuwait Airport Drone Attack; UNIFIL Peacekeeper Killed in Lebanon", "type": "military", "description": "Iranian-aligned militia drones strike Kuwait International Airport, damaging runway and terminal. One UNIFIL peacekeeper killed in southern Lebanon as Hezbollah-IDF fighting intensifies. Regional conflict spreading to previously neutral states.", "severity": 3, "fatalities": 3, "lat": 29.2266, "lon": 47.9689, "location": "Kuwait City, Kuwait"},
        {"date": "2026-03-26", "title": "Iran Rejects Direct US Talks; IRGC Opens Hormuz 'Tollbooth'", "type": "diplomatic", "description": "Iran formally rejects direct US negotiations, calling preconditions 'surrender terms.' IRGC confirms selective Hormuz transit for Chinese and Indian tankers paying transit fees — estimated $2-5M per passage. Western-flagged vessels still blocked. Brent rebounds to $108.01.", "severity": 2, "fatalities": 0, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-03-27", "title": "3,500 Marines Arrive on USS Tripoli; Humanitarian Ships Allowed", "type": "military", "description": "USS Tripoli arrives in Gulf of Oman with 3,500 Marines from 15th MEU. Pentagon announces humanitarian corridor through Hormuz for food and medical supplies. Iran allows passage of two humanitarian vessels as goodwill gesture while maintaining military blockade.", "severity": 2, "fatalities": 0, "lat": 25.0000, "lon": 57.5000, "location": "Gulf of Oman"},
        {"date": "2026-03-28", "title": "Houthis Fire Missiles at Israel; 82nd Airborne Deploys to Region", "type": "military", "description": "Houthis fire ballistic missiles at Israel for first time since war began — 11 injured from debris in southern Israel. Marks major Houthi escalation beyond Red Sea shipping attacks. 82nd Airborne Division deploys to region. Brent surges to $112.57. Renewed fears of Red Sea shipping strikes as Houthis re-enter direct combat.", "severity": 3, "fatalities": 0, "lat": 31.0461, "lon": 34.8516, "location": "Southern Israel"},
        {"date": "2026-03-29", "title": "Pentagon Prepares Ground Operations in Iran; 13 US KIA Total", "type": "military", "description": "Pentagon announces preparation for 'weeks of ground operations' targeting Qeshm Island and Kharg Island to permanently secure Hormuz and eliminate Iran's oil export capability. 82nd Airborne and Marine forces staging. Iran's IRGC threatens to 'rain fire' on any ground troops. US casualties now 13 KIA, 300+ wounded since Feb 28. Oil markets brace for further escalation.", "severity": 3, "fatalities": 0, "lat": 26.8500, "lon": 55.9000, "location": "Qeshm Island, Iran"},

        # ── Days 32-40 (March 31 – April 8, 2026) ──
        {"date": "2026-03-31", "title": "Continued Airstrikes Across Iran; Parliament Rejects Negotiations", "type": "military", "description": "US and Israeli air strikes continue with explosions reported in Tehran and Isfahan. Secretary Rubio says war objectives will be achieved in 'weeks, not months.' At least 4,700 Iranian security forces killed. Iranian Parliament Speaker Ghalibaf rejects negotiations, saying Iran cannot be forced into submission.", "severity": 3, "fatalities": 0, "lat": 35.6892, "lon": 51.3890, "location": "Tehran, Iran"},
        {"date": "2026-04-01", "title": "UK Announces 35-Nation Hormuz Meeting; Argentina Designates IRGC Terrorist Org", "type": "diplomatic", "description": "Britain announces meeting of ~35 countries to discuss reopening the Strait of Hormuz. China and Pakistan propose five-point ceasefire plan. Argentina designates IRGC as terrorist organization. Ships continue paying Iran tolls in yuan and cryptocurrency for Hormuz passage.", "severity": 1, "fatalities": 0, "lat": 51.5074, "lon": -0.1278, "location": "London, UK"},
        {"date": "2026-04-02", "title": "UK-Led 35-Nation Coalition Meets on Hormuz; Trump Says Objectives Nearly Met", "type": "diplomatic", "description": "UK Foreign Secretary Cooper chairs virtual meeting of ~35 countries to restore Hormuz freedom of navigation. US does not attend. PM Starmer calls for 'united front of military strength and diplomatic activity.' Trump states Washington close to achieving objectives. Hormuz evolves into dual-corridor system.", "severity": 1, "fatalities": 0, "lat": 51.5074, "lon": -0.1278, "location": "London, UK"},
        {"date": "2026-04-03", "title": "US F-15E Shot Down Over Iran; Search and Rescue Launched", "type": "military", "description": "American F-15E Strike Eagle shot down by Iranian forces. One crew member rescued alive, search continues for weapons system officer. An A-10 supporting the rescue is also struck; pilot ejects over Kuwait. Two UH-60 Black Hawks hit with minor injuries. Separately, US-Israel strikes a medical research center and bridge near Tehran.", "severity": 3, "fatalities": 1, "lat": 33.0000, "lon": 52.0000, "location": "Central Iran"},
        {"date": "2026-04-04", "title": "Projectile Strikes Near Bushehr Nuclear Plant; IAEA Issues Warning", "type": "nuclear", "description": "IAEA confirms projectile struck close to Bushehr nuclear power plant — fourth such incident. One protection staff member killed by fragment. No radiation increase detected. IAEA Director General Grossi: nuclear sites 'must never be attacked.' Russia evacuates 198 workers from facility.", "severity": 3, "fatalities": 1, "lat": 28.8333, "lon": 50.8833, "location": "Bushehr, Iran"},
        {"date": "2026-04-05", "title": "Trump Threatens Iran's Power Grid; Downed Pilot Rescued; Iraqi Militia Attacks US Embassy", "type": "military", "description": "Trump threatens to attack Iran's power plants and bridges if Hormuz not reopened within two days. Iranian-backed Iraqi militia Saraya Awliya al-Dam attacks US diplomatic facilities in Baghdad. Second F-15E crew member rescued alive. Iranian UAVs and cruise missiles intercepted in Qatari airspace.", "severity": 3, "fatalities": 0, "lat": 33.3152, "lon": 44.3661, "location": "Baghdad, Iraq"},
        {"date": "2026-04-06", "title": "Pakistan Proposes 'Islamabad Accord' — 45-Day Ceasefire Plan", "type": "diplomatic", "description": "Pakistan offers two-phased 45-day truce ('Islamabad Accord'). Pakistan army chief in contact with VP Vance, envoy Witkoff, and Iranian FM Araghchi. Iran 'positively reviewing' proposal but refuses to reopen Hormuz under temporary ceasefire. Houthis threaten to close Bab al-Mandeb strait — fears of second chokepoint crisis.", "severity": 2, "fatalities": 0, "lat": 33.6844, "lon": 73.0479, "location": "Islamabad, Pakistan"},
        {"date": "2026-04-07", "title": "Trump: 'A Whole Civilization Will Die Tonight'; Ceasefire Announced Before Deadline", "type": "diplomatic", "description": "Trump sets 8 PM ET deadline, posts 'a whole civilization will die tonight.' White House denies nuclear weapon plans. Pakistan PM makes last-ditch appeal. ~90 minutes before deadline, Trump announces two-week ceasefire based on Pakistan-mediated 10-point proposal. Iran to reopen Strait of Hormuz.", "severity": 2, "fatalities": 0, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},
        {"date": "2026-04-08", "title": "Islamabad Accords Ceasefire Takes Effect; Israel Launches 'Eternal Darkness' on Lebanon", "type": "diplomatic", "description": "Two-week ceasefire takes effect. Iran's Supreme National Security Council confirms acceptance. Hezbollah announces pause. But Israel launches 'Operation Eternal Darkness' — 100 airstrikes on Lebanon in 10 minutes targeting Hezbollah HQ and missile infrastructure. Gulf states intercept missiles; fires at Abu Dhabi's Habshan gas complex and Saudi pipeline. Hormuz remains effectively closed — 800+ freighters stuck. Brent crashes ~13% from pre-ceasefire highs.", "severity": 2, "fatalities": 0, "lat": 33.6844, "lon": 73.0479, "location": "Islamabad, Pakistan"},

        # ── Ceasefire Period (April 9-13, 2026) ──
        {"date": "2026-04-09", "title": "Ceasefire Violations Begin; Hormuz Still Closed; Only 4 Ships Transit", "type": "military", "description": "No sign of Hormuz blockade lifting — Iran limits crossings, charges tolls over $1M per ship. Only 4 dry cargo ships pass through. Iran accuses US/Israel of ceasefire violations over Lebanon strikes. IDF's 98th Division takes Bint Jbeil in southern Lebanon. Hezbollah fires rockets at Kiryat Shmona. 230 loaded oil tankers stuck inside the Gulf. Brent at $100.99.", "severity": 2, "fatalities": 0, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-04-10", "title": "Trump Accuses Iran of 'Very Poor Job' on Hormuz; Only 15 Ships Through", "type": "diplomatic", "description": "Trump accuses Iran of doing 'very poor job' managing Hormuz oil flow, says situation is 'not the agreement we have.' Only 15 ships total have made it through strait since ceasefire. Israel strikes Lebanon in early morning; Hezbollah launches rockets at Metula. Sirens in Tel Aviv, Haifa, and Ashdod. Ceasefire effectively collapsing. Brent at $97.78.", "severity": 2, "fatalities": 0, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},
        {"date": "2026-04-11", "title": "Vance Arrives in Islamabad for Peace Talks; US Navy Enters Hormuz", "type": "diplomatic", "description": "VP Vance, envoy Witkoff, and Jared Kushner arrive in Islamabad for talks with Iranian FM Araghchi and Parliament Speaker Ghalibaf. Several US Navy destroyers enter Strait of Hormuz for first time since war began — CENTCOM says mine clearance operations. Iran threatens to attack ships, accuses US of ceasefire violation. Trump announces US forces 'clearing' the strait.", "severity": 2, "fatalities": 0, "lat": 33.6844, "lon": 73.0479, "location": "Islamabad, Pakistan"},
        {"date": "2026-04-12", "title": "Islamabad Talks Collapse After 21 Hours; Trump Declares Naval Blockade", "type": "diplomatic", "description": "Vance leaves Pakistan after 21-hour marathon talks without agreement. Iranian FM spokesman says 'gaps on several major issues.' Ghalibaf says US 'did not succeed in gaining trust.' Trump threatens 'full naval blockade' — CENTCOM announces blockade starting Monday 10 AM EDT targeting vessels from/to Iranian ports. Non-Iran traffic free to transit. Brent at $96.", "severity": 2, "fatalities": 0, "lat": 33.6844, "lon": 73.0479, "location": "Islamabad, Pakistan"},
        {"date": "2026-04-13", "title": "US Naval Blockade Imminent; Oil Surges on Escalation Fears", "type": "military", "description": "Markets react to impending US naval blockade of Strait of Hormuz. Brent surges 6.95% to $103.72 as traders price in blockade risk. 500-700 vessels over 10,000 DWT stuck in Persian Gulf. Poll shows 63% of Israelis believe ceasefire does not extend to Lebanon. Ceasefire increasingly fragile ahead of Monday blockade deadline.", "severity": 2, "fatalities": 0, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},

        # ── Naval Blockade & Hormuz Reopening (April 14-17, 2026) ──
        {"date": "2026-04-14", "title": "US Naval Blockade Takes Effect; Trump Hints Talks Could Resume", "type": "military", "description": "US naval blockade of Iranian ports goes into force at 10 AM ET. CENTCOM says blockade applies to vessels entering/departing Iranian ports but 'will not impede freedom of navigation for vessels transiting the Strait of Hormuz to and from non-Iranian ports.' Trump hints US-Iran talks could resume within two days. White House confirms more peace deal talks in discussion. Brent at $100.19.", "severity": 2, "fatalities": 0, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
        {"date": "2026-04-15", "title": "Trump: Iran War 'Very Close to Over'; Israel-Lebanon Negotiations Advance", "type": "diplomatic", "description": "Day 47 of conflict. Trump says Iran war 'very close to over,' emphasizes Iran eager to make a deal and has 'agreed to things it hadn't two months ago.' Active Israel-Lebanon ceasefire negotiations underway. Brent drops to $96.83 on diplomatic optimism.", "severity": 1, "fatalities": 0, "lat": 38.9072, "lon": -77.0369, "location": "Washington, DC"},
        {"date": "2026-04-16", "title": "13 Iranian Oil Tankers Intercepted; Israel-Lebanon 10-Day Truce Announced", "type": "diplomatic", "description": "Gen. Dan Caine (JCS Chairman) announces 13 oil tankers intercepted enforcing Iran blockade. Trump announces Israel-Lebanon 10-day truce — separate from Iran negotiations. Trump: Israel 'prohibited' from bombing Lebanon during truce. Brent ends at $94.89 as Lebanon ceasefire eases regional tensions.", "severity": 1, "fatalities": 0, "lat": 33.8938, "lon": 35.5018, "location": "Beirut, Lebanon"},
        {"date": "2026-04-17", "title": "STRAIT OF HORMUZ REOPENED FOR COMMERCIAL SHIPPING", "type": "diplomatic", "description": "Iran FM Araghchi declares Strait of Hormuz 'completely open for commercial ships for remainder of ceasefire' — explicitly tied to Lebanon 10-day truce duration. US blockade of Iranian ports remains in full force pending peace deal. Starmer-Trump call on assembling diplomatic coalition for long-term Hormuz reopening. France/UK rally 40+ nations on defensive plan. Ceasefire continues. Partial breakthrough but full normalization contingent on broader peace deal.", "severity": 1, "fatalities": 0, "lat": 26.5667, "lon": 56.2500, "location": "Strait of Hormuz"},
    ]


def fetch_iran_news() -> List[dict]:
    """Fetch live Iran/oil war news headlines from Google News RSS. No API key needed.

    Cache TTL is 2 hours but we also serve stale cache as the FINAL fallback
    when the live fetch fails — better to show yesterday's headlines than
    a blank feed (the 3-second per-RSS timeout means transient network blips
    used to silently empty the news ticker).
    """
    cached = _read_cache("iran_news", 7200)  # 2-hour cache
    if cached:
        logger.info("Iran news: serving from cache")
        return cached

    try:
        import xml.etree.ElementTree as ET

        # Multiple targeted queries to capture different angles of the conflict
        queries = [
            "Iran+war+US+strikes",
            "Strait+of+Hormuz+oil+shipping",
            "Brent+crude+oil+Iran",
        ]
        seen_titles = set()
        all_articles = []

        def _fetch_rss(q):
            url = f"https://news.google.com/rss/search?q={q}&hl=en-US&gl=US&ceid=US:en"
            resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=3)
            if resp.status_code != 200:
                return []
            root = ET.fromstring(resp.text)
            items = []
            for item in root.findall(".//item"):
                title_el = item.find("title")
                pub_el = item.find("pubDate")
                source_el = item.find("source")
                link_el = item.find("link")
                if title_el is None or pub_el is None:
                    continue
                items.append({
                    "title": title_el.text or "",
                    "pubDate": pub_el.text or "",
                    "source": source_el.text if source_el is not None else "",
                    "url": link_el.text if link_el is not None else "",
                })
            return items

        # Fetch all RSS feeds in parallel
        with ThreadPoolExecutor(max_workers=3) as pool:
            rss_results = list(pool.map(_fetch_rss, queries))

        for items in rss_results:
            for a in items:
                title_lower = a["title"].lower()
                if title_lower not in seen_titles:
                    seen_titles.add(title_lower)
                    all_articles.append(a)

        # Relevance filter: must mention Iran/Hormuz AND oil/military/war context
        iran_terms = {"iran", "iranian", "tehran", "hormuz", "irgc", "persian gulf", "hezbollah", "houthi"}
        context_terms = {"oil", "brent", "crude", "war", "strike", "military", "missile",
                         "drone", "attack", "bomb", "navy", "sanctions", "nuclear",
                         "ceasefire", "tanker", "shipping", "casualt", "killed", "troops"}

        filtered = []
        for a in all_articles:
            t = a["title"].lower()
            has_iran = any(term in t for term in iran_terms)
            has_context = any(term in t for term in context_terms)
            if has_iran and has_context:
                filtered.append(a)

        # Classify type based on keywords
        def classify_type(title):
            t = title.lower()
            if any(w in t for w in ["oil", "brent", "crude", "price", "opec", "tanker", "shipping", "hormuz", "trade"]):
                return "proxy"
            if any(w in t for w in ["sanction", "embargo", "treasury", "ofac"]):
                return "sanctions"
            if any(w in t for w in ["nuclear", "uranium", "iaea", "enrich"]):
                return "nuclear"
            if any(w in t for w in ["talk", "negotiat", "diplomat", "ceasefire", "peace", "deal", "summit"]):
                return "diplomatic"
            return "military"

        # Parse dates and build structured output
        results = []
        for a in filtered[:50]:  # Cap at 50 most recent
            try:
                pub_dt = datetime.strptime(a["pubDate"][:25], "%a, %d %b %Y %H:%M:%S")
                date_str = pub_dt.strftime("%Y-%m-%d")
            except (ValueError, IndexError):
                continue

            results.append({
                "date": date_str,
                "title": a["title"],
                "type": classify_type(a["title"]),
                "source": a["source"],
                "url": a["url"],
                "severity": 3,  # Default; news headlines don't have severity
                "auto": True,   # Flag to distinguish from curated events
            })

        # Sort by date descending
        results.sort(key=lambda x: x["date"], reverse=True)

        if results:
            _write_cache("iran_news", results)
            logger.info(f"Iran news: fetched {len(results)} relevant articles from Google News")
        return results

    except Exception as e:
        logger.warning(f"Iran news fetch failed: {e}")

    # Final fallback: serve the most recent cached headlines even past TTL.
    # Without this, a single RSS hiccup wipes the news ticker until the next
    # successful fetch — empty UI is worse than a 3-hour-old headline.
    stale = _read_stale_cache("iran_news") or []
    if stale:
        logger.info(f"Iran news: serving stale cache ({len(stale)} items)")
    return stale


# ─── IMF PortWatch — Chokepoint Transit Data ─────────────────────────────────

_PORTWATCH_BASE = (
    "https://services9.arcgis.com/weJ1QsnbMYJlCHdG/arcgis/rest/services/"
    "Daily_Chokepoints_Data/FeatureServer/0/query"
)

_CHOKEPOINT_IDS = {
    "suez": "chokepoint1",
    "bab_el_mandeb": "chokepoint4",
    "hormuz": "chokepoint6",
}


def _fetch_chokepoint_transits(chokepoint_key: str) -> List[dict]:
    """Fetch daily transit data from IMF PortWatch for a given chokepoint and aggregate to monthly totals.
    Returns list of {month, transits, tanker_transits, container_transits} sorted ascending."""

    cache_key = f"{chokepoint_key}_transits"
    cached = _read_cache(cache_key, ttl=86400)  # 24h cache — data updates weekly
    if cached:
        return cached

    port_id = _CHOKEPOINT_IDS.get(chokepoint_key)
    if not port_id:
        logger.error(f"Unknown chokepoint key: {chokepoint_key}")
        return []

    logger.info(f"Fetching {chokepoint_key} transit data from IMF PortWatch...")

    all_records = []
    offset = 0
    batch_size = 2000

    while True:
        params = {
            "where": f"portid='{port_id}' AND year>=2023 AND (year>2023 OR month>=7)",
            "outFields": "date,year,month,n_total,n_tanker,n_container,n_dry_bulk",
            "orderByFields": "date ASC",
            "resultRecordCount": batch_size,
            "resultOffset": offset,
            "f": "json",
        }
        try:
            resp = requests.get(_PORTWATCH_BASE, params=params, headers=_BROWSER_HEADERS, timeout=10)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            logger.warning(f"PortWatch API request failed for {chokepoint_key} at offset {offset}: {e}")
            break

        features = data.get("features", [])
        if not features:
            break

        for f in features:
            all_records.append(f["attributes"])

        if len(features) < batch_size:
            break
        offset += batch_size

    if not all_records:
        logger.warning(f"No {chokepoint_key} transit data returned from PortWatch")
        return []

    # Aggregate daily records to monthly totals
    monthly = defaultdict(lambda: {"transits": 0, "tanker": 0, "container": 0, "dry_bulk": 0, "days": 0})
    for r in all_records:
        key = f"{r['year']}-{r['month']:02d}"
        monthly[key]["transits"] += r.get("n_total", 0) or 0
        monthly[key]["tanker"] += r.get("n_tanker", 0) or 0
        monthly[key]["container"] += r.get("n_container", 0) or 0
        monthly[key]["dry_bulk"] += r.get("n_dry_bulk", 0) or 0
        monthly[key]["days"] += 1

    # Build sorted result — only include months with at least 15 days of data
    result = []
    for month_key in sorted(monthly.keys()):
        m = monthly[month_key]
        if m["days"] < 15:
            continue
        result.append({
            "month": month_key,
            "transits": m["transits"],
            "tanker_transits": m["tanker"],
            "container_transits": m["container"],
            "dry_bulk_transits": m["dry_bulk"],
            "days_sampled": m["days"],
        })

    logger.info(f"{chokepoint_key} transit data: {len(result)} months from {len(all_records)} daily records")
    _write_cache(cache_key, result)
    return result


def fetch_suez_transits() -> List[dict]:
    """Fetch Suez Canal transit data."""
    return _fetch_chokepoint_transits("suez")


def fetch_bab_el_mandeb_transits() -> List[dict]:
    """Fetch Bab el-Mandeb Strait transit data."""
    return _fetch_chokepoint_transits("bab_el_mandeb")


def fetch_hormuz_transits() -> List[dict]:
    """Fetch Strait of Hormuz transit data."""
    return _fetch_chokepoint_transits("hormuz")


_HIGH_SIGNAL_WORDS = {"killed", "strike", "strikes", "missile", "attack", "destroyed", "explosion", "bomb", "drone", "fire", "burning", "casualties", "dead", "wounded"}
_ACLED_KEY_ACTORS = {"irgc", "united states", "israel", "hezbollah", "houthi", "navy", "air force", "centcom", "idf"}
_ACLED_TYPE_MAP = {
    "Battles": "military",
    "Explosions/Remote violence": "military",
    "Violence against civilians": "military",
    "Strategic developments": "diplomatic",
    "Protests": "proxy",
    "Riots": "proxy",
}

# Words that indicate a roundup/summary headline rather than a discrete event
_ROUNDUP_WORDS = {"live updates", "live:", "latest:", "what we know", "what happened",
                   "day of war", "here's what", "key developments", "live update",
                   "minute by minute", "as it happened", "rolling coverage"}

# High-credibility sources get a scoring boost
_TIER1_SOURCES = {"reuters", "associated press", "ap news", "bbc", "the new york times",
                  "the washington post", "the wall street journal", "al jazeera",
                  "the guardian", "cnn", "abc news", "nbc news", "financial times"}

# Words that indicate a major discrete event (strong signal)
_MAJOR_EVENT_WORDS = {
    "seize": 3, "seized": 3, "sinks": 4, "sunk": 4, "torpedoed": 4,
    "invades": 4, "invasion": 4, "ceasefire": 4, "surrenders": 4,
    "assassinated": 4, "killed": 3, "destroys": 3, "destroyed": 3,
    "launches": 2, "fires": 2, "strikes": 2, "hits": 2, "shoots": 2,
    "intercepts": 2, "blocks": 2, "closes": 3, "shuts": 2,
    "deploys": 2, "mobilizes": 2, "evacuates": 2,
    "sanctions": 2, "ultimatum": 3, "declares": 2, "threatens": 1,
    "surges": 2, "crashes": 2, "spikes": 2, "plunges": 2,
    "mined": 3, "mines": 2, "torpedo": 3, "boarding": 2,
    "shoots down": 3, "shot down": 3,
}


def _score_news_headline(item: dict) -> int:
    """Score a news headline for significance. Higher = more important."""
    title = item.get("title", "")
    title_lower = title.lower()
    source = item.get("source", "").lower()
    score = 0

    # Reject roundup/summary headlines — these aren't discrete events
    if any(rw in title_lower for rw in _ROUNDUP_WORDS):
        return -1

    # Major event keyword scoring
    for word, pts in _MAJOR_EVENT_WORDS.items():
        if word in title_lower:
            score += pts

    # High-signal words (general)
    words = set(title_lower.split())
    score += len(words & _HIGH_SIGNAL_WORDS) * 2

    # Named key actors
    for actor in _ACLED_KEY_ACTORS:
        if actor in title_lower:
            score += 1

    # Specificity bonus: numbers in headline (casualties, counts, distances)
    if re.search(r'\b\d+\b', title):
        score += 1

    # Source credibility boost
    if any(s in source for s in _TIER1_SOURCES):
        score += 2

    # Location specificity bonus (not just "Iran" — a specific place)
    specific_locations = {"hormuz", "natanz", "isfahan", "fordow", "bushehr", "bandar abbas",
                          "tehran", "qeshm", "fujairah", "ras tanura", "salalah", "jask"}
    if any(loc in title_lower for loc in specific_locations):
        score += 2

    return score


def _geocode_news_events(news_items: List[dict]) -> List[dict]:
    """Convert significant Google News headlines into curated-quality timeline events.

    Scores headlines for significance and promotes only the top 1-2 per day,
    filtering out roundup articles and low-signal noise.
    """
    # Score and geocode all candidates
    candidates = []
    for item in news_items:
        title = item.get("title", "")
        title_lower = title.lower()

        score = _score_news_headline(item)
        if score < 4:  # Minimum threshold — must be clearly significant
            continue

        # Find location by longest-first keyword match
        lat, lon, location = None, None, None
        for key in _LOCATION_KEYS_SORTED:
            if key in title_lower:
                lat, lon, location = CONFLICT_THEATER_LOCATIONS[key]
                break

        if lat is None:
            continue  # Can't plot without coordinates

        severity = 5 if score >= 10 else (4 if score >= 6 else 3)

        candidates.append({
            "date": item.get("date", "")[:10],
            "title": title,
            "type": item.get("type", "military"),
            "description": f"Live news via {item.get('source', 'Unknown')}",
            "severity": severity,
            "lat": lat,
            "lon": lon,
            "location": location,
            "fatalities": 0,
            "source_type": "news_auto",
            "_score": score,
        })

    # Keep only top 2 per date (highest scoring)
    by_date = defaultdict(list)
    for c in candidates:
        by_date[c["date"]].append(c)

    results = []
    for date, events in by_date.items():
        events.sort(key=lambda x: x["_score"], reverse=True)
        for ev in events[:2]:
            del ev["_score"]
            results.append(ev)

    logger.info(f"News scoring: {len(news_items)} articles → {len(candidates)} significant → {len(results)} promoted")
    return results


def _promote_acled_events(acled_events: List[dict], curated_events: List[dict]) -> List[dict]:
    """Score and promote the most significant ACLED events to curated quality."""
    # Build curated index for deduplication: (date, lat, lon)
    curated_index = []
    for c in curated_events:
        try:
            curated_index.append((c["date"], float(c.get("lat", 0)), float(c.get("lon", 0))))
        except (ValueError, TypeError):
            pass

    def _is_near_curated(date_str, lat, lon):
        for c_date, c_lat, c_lon in curated_index:
            if abs(ord(date_str[8]) - ord(c_date[8])) <= 1 and date_str[:7] == c_date[:7]:
                if abs(lat - c_lat) < 0.5 and abs(lon - c_lon) < 0.5:
                    return True
        return False

    scored = []
    for e in acled_events:
        try:
            lat = float(e.get("latitude", 0))
            lon = float(e.get("longitude", 0))
        except (ValueError, TypeError):
            continue
        if not lat or not lon:
            continue

        fatalities = int(e.get("fatalities", 0) or 0)
        event_type = e.get("event_type", "")
        sub_type = e.get("sub_event_type", "")
        actor1 = (e.get("actor1", "") or "").lower()
        actor2 = (e.get("actor2", "") or "").lower()
        location = e.get("location", "")
        date_str = (e.get("event_date", "") or "")[:10]

        if not date_str or len(date_str) < 10:
            continue

        # Skip if near an existing curated event
        if _is_near_curated(date_str, lat, lon):
            continue

        # Skip domestic protests/riots — not war-relevant
        if event_type in ("Protests", "Riots"):
            continue

        # Score
        score = 0
        score += min(30, fatalities * 3)
        if event_type in ("Explosions/Remote violence", "Battles"):
            score += 5
        if any(k in actor1 or k in actor2 for k in _ACLED_KEY_ACTORS):
            score += 5
        loc_lower = location.lower()
        if any(k in loc_lower for k in _LOCATION_KEYS_SORTED[:20]):
            score += 3
        if any(k in sub_type.lower() for k in ("air/drone strike", "shelling", "artillery", "armed clash", "suicide bomb")):
            score += 2

        if score < 8:
            continue

        # Generate title
        if sub_type and location:
            title = f"{sub_type} in {location}"
        elif actor1 and location:
            actor_display = actor1.split("(")[0].strip().title()
            title = f"{event_type}: {actor_display} in {location}"
        else:
            title = f"{event_type} in {location or 'Iran'}"

        # Map severity
        if score >= 19:
            severity = 5
        elif score >= 13:
            severity = 4
        else:
            severity = 3

        scored.append({
            "date": date_str,
            "title": title,
            "type": _ACLED_TYPE_MAP.get(event_type, "military"),
            "description": (e.get("notes", "") or "")[:500],
            "severity": severity,
            "lat": lat,
            "lon": lon,
            "location": location,
            "fatalities": fatalities,
            "source_type": "acled_promoted",
            "_score": score,
        })

    # Sort by score, take top 20
    scored.sort(key=lambda x: x["_score"], reverse=True)
    for item in scored:
        del item["_score"]
    return scored[:20]


def get_merged_curated_events(include_news: bool = True) -> List[dict]:
    """Merge hardcoded curated events with auto-discovered OSINT events.

    Args:
        include_news: If True, geocode live news headlines and merge.
                      Set False when called from compute_iran_impact() to avoid
                      blocking on Google News fetches.
    """
    # 1. Hardcoded curated events
    curated = get_curated_iran_events()
    for e in curated:
        if "source_type" not in e:
            e["source_type"] = "curated"

    news_events = []

    # 2. Geocoded news events (only if requested and cache is warm)
    if include_news:
        try:
            # Only use cached news — never trigger a live fetch here
            cached_news = _read_cache("iran_news", 7200)
            if cached_news:
                news_events = _geocode_news_events(cached_news)
                # Deduplicate news against curated (same date + nearby location)
                curated_index = [(c["date"], c.get("lat", 0), c.get("lon", 0)) for c in curated]
                deduped_news = []
                for ne in news_events:
                    is_dup = False
                    for c_date, c_lat, c_lon in curated_index:
                        if ne["date"] == c_date and abs(ne["lat"] - c_lat) < 0.5 and abs(ne["lon"] - c_lon) < 0.5:
                            is_dup = True
                            break
                    if not is_dup:
                        deduped_news.append(ne)
                news_events = deduped_news
        except Exception as e:
            logger.warning(f"News geocoding failed: {e}")
            news_events = []

    # 3. Merge and sort (curated + news only; ACLED-promoted are internal Iranian
    #    conflicts like armed clashes/IEDs, not US-Iran tensions — excluded)
    merged = curated + news_events
    merged.sort(key=lambda x: x.get("date", ""), reverse=True)

    logger.info(f"Merged events: {len(curated)} curated + {len(news_events)} news = {len(merged)} total")
    return merged


def compute_iran_impact(iran_events: list, brent_prices: list) -> dict:
    """Calculate oil price impact metrics around Iran events."""
    if not brent_prices:
        return {"kpis": {}, "impact_by_type": {}, "event_table": []}

    # Build price lookup by date
    price_map = {p["date"]: p["price"] for p in brent_prices}
    sorted_dates = sorted(price_map.keys())

    def get_closing_price_before(date_str: str):
        """Get the closing price on the most recent trading day BEFORE this date."""
        idx = bisect.bisect_left(sorted_dates, date_str)
        if idx > 0:
            return price_map[sorted_dates[idx - 1]]
        return None

    def get_closing_price_after(date_str: str, days: int):
        """Get the closing price on the nearest trading day on or after date + days."""
        target = (datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=days)).strftime("%Y-%m-%d")
        idx = bisect.bisect_left(sorted_dates, target)
        if idx < len(sorted_dates):
            return price_map[sorted_dates[idx]]
        # If past the end, use the last available price
        if sorted_dates:
            return price_map[sorted_dates[-1]]
        return None

    # Get curated events for impact table (skip news fetch — avoid blocking)
    curated = get_merged_curated_events(include_news=False)

    # Group curated events by TRADING day.
    # Weekend events (Sat/Sun) roll forward to the next Monday when markets
    # actually reacted, so before/after prices form a clean chain.

    def _next_trading_day(date_str: str) -> str:
        """If date falls on a weekend, roll forward to Monday."""
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        while dt.weekday() >= 5:  # 5=Sat, 6=Sun
            dt += timedelta(days=1)
        return dt.strftime("%Y-%m-%d")

    events_by_trading_day = OrderedDict()
    for ev in curated:
        td = _next_trading_day(ev["date"])
        if td not in events_by_trading_day:
            events_by_trading_day[td] = []
        events_by_trading_day[td].append(ev)

    # Build impact table — one row per trading day
    event_table = []
    for td, day_events in events_by_trading_day.items():
        price_before = get_closing_price_before(td)
        price_after = get_closing_price_after(td, 0)  # Trading-day close

        change_pct = None
        if price_before and price_after:
            change_pct = round((price_after - price_before) / price_before * 100, 2)

        # Collect all event details for this trading day
        max_severity = max(ev["severity"] for ev in day_events)
        types = list(dict.fromkeys(ev["type"] for ev in day_events))  # unique, order-preserving

        # Show the actual date range if events span multiple calendar days
        event_dates = sorted(set(ev["date"] for ev in day_events))
        display_date = event_dates[0] if len(event_dates) == 1 else f"{event_dates[0]} – {event_dates[-1]}"

        event_table.append({
            "date": td,                        # trading day (for sorting/filtering)
            "display_date": display_date,       # shows date range if weekend events rolled in
            "events": [{"title": ev["title"], "type": ev["type"], "severity": ev["severity"],
                         "description": ev.get("description", ""),
                         "actual_date": ev["date"]} for ev in day_events],
            "title": day_events[0]["title"],    # primary event title (backwards compat)
            "type": types[0],                   # primary type (backwards compat)
            "types": types,                     # all types for the day
            "severity": max_severity,
            "brent_before": round(price_before, 2) if price_before else None,
            "brent_after": round(price_after, 2) if price_after else None,
            "change_pct": change_pct,
        })

    # Aggregate impact by event type at different horizons
    curated_by_type = {}
    for ev in curated:
        t = ev["type"]
        if t not in curated_by_type:
            curated_by_type[t] = []
        curated_by_type[t].append(ev["date"])

    impact_by_type = {}
    for etype, dates in curated_by_type.items():
        offsets = {"T+1": 1, "T+3": 3, "T+5": 5, "T+7": 7}
        type_impact = {}
        for label, off in offsets.items():
            changes = []
            for d in dates:
                pb = get_closing_price_before(d)
                pa = get_closing_price_after(d, off)
                if pb and pa:
                    changes.append((pa - pb) / pb * 100)
            type_impact[label] = round(sum(changes) / len(changes), 3) if changes else 0
        impact_by_type[etype] = type_impact

    # ACLED-based aggregation
    acled_dates = list({e.get("event_date", "")[:10] for e in iran_events if e.get("event_date")})
    all_changes_3d = []
    max_vol_spike = 0
    current_month = datetime.now().strftime("%Y-%m")
    acled_this_month = sum(1 for e in iran_events if (e.get("event_date") or "").startswith(current_month))
    curated_this_month = sum(1 for e in curated if e["date"].startswith(current_month))
    events_this_month = acled_this_month + curated_this_month

    for d in acled_dates:
        pb = get_closing_price_before(d)
        pa = get_closing_price_after(d, 3)
        if pb and pa:
            change = abs(pa - pb)
            all_changes_3d.append(change)
            if change > max_vol_spike:
                max_vol_spike = change

    kpis = {
        "total_events": len(iran_events),
        "avg_price_move_3d": round(sum(all_changes_3d) / len(all_changes_3d), 2) if all_changes_3d else 0,
        "peak_volatility_spike": round(max_vol_spike, 2),
        "events_this_month": events_this_month,
    }

    return {
        "kpis": kpis,
        "impact_by_type": impact_by_type,
        "event_table": event_table,
    }


# ─── Iran Conflict Intensity Index ─────────────────────────────────────────────

def compute_iran_intensity() -> List[dict]:
    """Build a daily/weekly Iran conflict intensity index from ACLED + curated events.
    Analogous to WeeklyAttackFreq from the Houthi thesis analysis.

    Scoring:
      - ACLED events: Battles/Explosions=3, Violence against civilians=2, other=1
      - Curated events: weighted by severity (1-5)
      - Fatalities: log-scaled bonus
      - Hormuz proximity (<200km): 2x multiplier
    """
    # Get data sources
    iran_acled = _read_cache("iran_events", 7200) or _load_iran_json_fallback() or []
    curated = get_curated_iran_events()

    # ACLED event type weights
    acled_weights = {
        "Battles": 3, "Explosions/Remote violence": 3,
        "Violence against civilians": 2, "Strategic developments": 1,
        "Protests": 1, "Riots": 1,
    }

    # Hormuz coordinates for proximity check
    HORMUZ_LAT, HORMUZ_LON = 26.5667, 56.2500

    def _haversine_km(lat1, lon1, lat2, lon2):
        R = 6371
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
        return R * 2 * math.asin(math.sqrt(a))

    # Build daily scores
    daily_scores = {}

    # Score ACLED events
    for e in iran_acled:
        date = (e.get("event_date") or "")[:10]
        if not date or date < "2025-01-01":
            continue
        etype = e.get("event_type", "")
        weight = acled_weights.get(etype, 1)
        fatalities = int(e.get("fatalities", 0) or 0)
        fat_bonus = math.log1p(fatalities) if fatalities > 0 else 0

        # Proximity multiplier
        try:
            lat = float(e.get("latitude", 0))
            lon = float(e.get("longitude", 0))
            dist = _haversine_km(lat, lon, HORMUZ_LAT, HORMUZ_LON)
            proximity = 2.0 if dist < 200 else 1.0
        except (ValueError, TypeError):
            proximity = 1.0

        score = (weight + fat_bonus) * proximity
        daily_scores[date] = daily_scores.get(date, 0) + score

    # Score curated events (higher weight — these are confirmed major events)
    for e in curated:
        date = e["date"]
        severity = e.get("severity", 1)
        fatalities = e.get("fatalities", 0) or 0
        fat_bonus = math.log1p(fatalities) if fatalities > 0 else 0

        try:
            lat = float(e.get("lat", 0))
            lon = float(e.get("lon", 0))
            dist = _haversine_km(lat, lon, HORMUZ_LAT, HORMUZ_LON)
            proximity = 2.0 if dist < 200 else 1.0
        except (ValueError, TypeError):
            proximity = 1.0

        score = (severity * 2 + fat_bonus) * proximity
        daily_scores[date] = daily_scores.get(date, 0) + score

    # Build time series with 7-day rolling sum
    if not daily_scores:
        return []

    sorted_dates = sorted(daily_scores.keys())
    start = datetime.strptime(sorted_dates[0], "%Y-%m-%d")
    end = datetime.strptime(sorted_dates[-1], "%Y-%m-%d")

    result = []
    current = start
    window = []
    while current <= end:
        ds = current.strftime("%Y-%m-%d")
        daily = round(daily_scores.get(ds, 0), 2)
        window.append(daily)
        if len(window) > 7:
            window.pop(0)
        weekly = round(sum(window), 2)
        result.append({"date": ds, "daily_intensity": daily, "weekly_intensity": weekly})
        current += timedelta(days=1)

    return result


def compute_hormuz_disruption() -> dict:
    """Compare current Hormuz transit rates to pre-war baseline."""
    hormuz = _read_cache("hormuz_transits", 86400)
    if not hormuz:
        try:
            hormuz = _fetch_chokepoint_transits("hormuz")
        except Exception:
            hormuz = []

    if not hormuz or len(hormuz) < 3:
        return {"status": "UNKNOWN", "baseline_transits": None, "current_transits": None, "pct_decline": None}

    # Pre-war baseline: average of months before Feb 2026
    pre_war = [m for m in hormuz if m["month"] < "2026-02"]
    current = [m for m in hormuz if m["month"] >= "2026-02"]

    if not pre_war or not current:
        return {"status": "UNKNOWN", "baseline_transits": None, "current_transits": None, "pct_decline": None}

    baseline_avg = sum(m["transits"] for m in pre_war) / len(pre_war)
    current_avg = sum(m["transits"] for m in current) / len(current)
    pct_decline = round((1 - current_avg / baseline_avg) * 100, 1) if baseline_avg > 0 else 0

    # Tanker-specific if available
    tanker_baseline = None
    tanker_current = None
    tanker_decline = None
    if pre_war[0].get("tanker_transits") is not None:
        tanker_baseline = sum(m.get("tanker_transits", 0) for m in pre_war) / len(pre_war)
        tanker_current = sum(m.get("tanker_transits", 0) for m in current) / len(current)
        tanker_decline = round((1 - tanker_current / tanker_baseline) * 100, 1) if tanker_baseline > 0 else 0

    # Status classification — PortWatch data lags. Override with current events.
    today = datetime.now().strftime("%Y-%m-%d")

    # As of Apr 17, 2026: Iran declared Hormuz open for remainder of Lebanon ceasefire
    status_note = None
    if today >= "2026-04-17":
        status = "CONDITIONAL"
        status_note = "Iran declared Hormuz 'completely open for commercial ships' on Apr 17 (tied to 10-day Lebanon ceasefire). US blockade of Iranian ports remains in force pending peace deal."
    elif today >= "2026-04-13":
        status = "BLOCKADE"
        status_note = "US naval blockade enforcing against Iranian ports (non-Iran traffic permitted in strait)."
    elif pct_decline >= 80:
        status = "BLOCKADE"
    elif pct_decline >= 40:
        status = "RESTRICTED"
    elif pct_decline >= 10:
        status = "DISRUPTED"
    else:
        status = "OPEN"

    return {
        "status": status,
        "status_note": status_note,
        "baseline_transits": round(baseline_avg),
        "current_transits": round(current_avg),
        "pct_decline": pct_decline,
        "tanker_baseline": round(tanker_baseline) if tanker_baseline else None,
        "tanker_current": round(tanker_current) if tanker_current else None,
        "tanker_pct_decline": tanker_decline,
        "months_data": hormuz,
    }


def get_war_phases() -> List[dict]:
    """Return war phase definitions for timeline annotations."""
    return [
        {"phase": 1, "name": "Maximum Pressure Restored", "start": "2025-01-20", "end": "2026-02-27", "color": "#F09060"},
        {"phase": 2, "name": "Twelve-Day War", "start": "2026-02-28", "end": "2026-03-11", "color": "#E05555"},
        {"phase": 3, "name": "Hormuz Blockade & Escalation", "start": "2026-03-12", "end": "2026-03-19", "color": "#9B8EC4"},
        {"phase": 4, "name": "Regional Spread", "start": "2026-03-20", "end": "2026-03-29", "color": "#D4B870"},
        {"phase": 5, "name": "Ground Operations Prep", "start": "2026-03-30", "end": "2026-04-07", "color": "#E05555"},
        {"phase": 6, "name": "Ceasefire & Islamabad Talks", "start": "2026-04-08", "end": "2026-04-12", "color": "#4A90D9"},
        {"phase": 7, "name": "Naval Blockade", "start": "2026-04-13", "end": "2026-04-16", "color": "#E05555"},
        {"phase": 8, "name": "Hormuz Reopened (Conditional)", "start": "2026-04-17", "end": "2026-12-31", "color": "#00e676"},
    ]


def get_comparative_data() -> dict:
    """Build Houthi vs Iran comparative analysis dataset."""
    # Houthi/baseline timeseries — pulled from load_master_dataset() so it
    # picks up the live-extended rows. Reading the raw CSV directly (as this
    # function used to) freezes the timeseries at whatever date the dev last
    # regenerated master_dataset.csv, which left charts dead-ending months
    # before wall-clock.
    houthi_ts = []
    try:
        master = load_master_dataset()
        for row in (master.get("timeseries") or []):
            houthi_ts.append({
                "date": (row.get("date") or "")[:10],
                "brent_price": row.get("brent_price"),
                "weekly_attacks": row.get("weekly_attacks"),
                "daily_volatility": row.get("daily_volatility"),
            })
    except Exception as e:
        logger.warning(f"Comparative: master timeseries load failed: {e}")

    # Iran intensity data
    iran_intensity = compute_iran_intensity()

    # Brent prices (includes war period)
    brent = _read_cache("brent_prices", 7200) or []

    # Hormuz disruption
    hormuz = compute_hormuz_disruption()

    # Chokepoint transit comparison
    bab_transits = _read_cache("bab_el_mandeb_transits", 86400) or []
    hormuz_transits = hormuz.get("months_data", [])

    # War phases
    phases = get_war_phases()

    # Houthi period stats
    houthi_prices = [h["brent_price"] for h in houthi_ts if h["brent_price"]]
    houthi_vol = [h["daily_volatility"] for h in houthi_ts if h["daily_volatility"]]

    # Iran period stats (from curated events + Brent)
    iran_brent = [b for b in brent if b["date"] >= "2026-02-28"]
    iran_prices = [b["price"] for b in iran_brent]

    return {
        "houthi_timeseries": houthi_ts,
        "iran_intensity": iran_intensity,
        "iran_brent": iran_brent,
        "bab_transits": bab_transits,
        "hormuz_transits": hormuz_transits,
        "hormuz_disruption": hormuz,
        "war_phases": phases,
        "summary": {
            "houthi_period": "Oct 2023 – Oct 2025",
            "houthi_price_range": [round(min(houthi_prices), 2), round(max(houthi_prices), 2)] if houthi_prices else None,
            "houthi_avg_volatility": round(sum(houthi_vol) / len(houthi_vol), 4) if houthi_vol else None,
            "iran_period": "Feb 28, 2026 – Present",
            "iran_price_range": [round(min(iran_prices), 2), round(max(iran_prices), 2)] if iran_prices else None,
            "iran_price_change_pct": round((iran_prices[-1] - iran_prices[0]) / iran_prices[0] * 100, 1) if len(iran_prices) >= 2 else None,
            "thesis_finding": "Houthi attacks had minimal impact on oil volatility (market adaptation). Iran war caused massive price disruption — state-actor chokepoint closure breaks the adaptation model.",
        },
    }


def _t_cdf_approx(t: float, df: int) -> float:
    """Approximate Student's t-distribution CDF using a series expansion.
    Good enough for p-value calculation without requiring scipy.

    For df > 30, normal approximation is very accurate.
    For smaller df, uses the relation to the incomplete beta function
    via a continued fraction approximation.
    """
    if df <= 0:
        return 0.5

    # For large df, use normal approximation via erf
    if df > 30:
        return 0.5 * (1 + math.erf(t / math.sqrt(2)))

    # For smaller df, use the relation: P(T <= t) = 1 - 0.5 * I(df/(df+t^2); df/2, 0.5)
    # where I is the regularized incomplete beta function.
    # We use an approximation from Abramowitz & Stegun.
    x = df / (df + t * t)

    # Regularized incomplete beta via continued fraction (Lentz's method)
    def _incbeta(a, b, x):
        if x <= 0:
            return 0.0
        if x >= 1:
            return 1.0
        # Use continued fraction
        bt = math.exp(math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b) +
                      a * math.log(x) + b * math.log(1 - x))
        if x < (a + 1) / (a + b + 2):
            return bt * _betacf(a, b, x) / a
        else:
            return 1 - bt * _betacf(b, a, 1 - x) / b

    def _betacf(a, b, x):
        MAXIT = 200
        EPS = 3e-7
        qab = a + b
        qap = a + 1
        qam = a - 1
        c = 1.0
        d = 1.0 - qab * x / qap
        if abs(d) < 1e-30:
            d = 1e-30
        d = 1.0 / d
        h = d
        for m in range(1, MAXIT + 1):
            m2 = 2 * m
            aa = m * (b - m) * x / ((qam + m2) * (a + m2))
            d = 1.0 + aa * d
            if abs(d) < 1e-30:
                d = 1e-30
            c = 1.0 + aa / c
            if abs(c) < 1e-30:
                c = 1e-30
            d = 1.0 / d
            h *= d * c
            aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
            d = 1.0 + aa * d
            if abs(d) < 1e-30:
                d = 1e-30
            c = 1.0 + aa / c
            if abs(c) < 1e-30:
                c = 1e-30
            d = 1.0 / d
            dl = d * c
            h *= dl
            if abs(dl - 1.0) < EPS:
                break
        return h

    try:
        ib = _incbeta(df / 2, 0.5, x)
        if t >= 0:
            return 1 - 0.5 * ib
        else:
            return 0.5 * ib
    except Exception:
        # Fallback to normal approximation
        return 0.5 * (1 + math.erf(t / math.sqrt(2)))


def _two_sided_pvalue(t: float, df: int) -> float:
    """Two-sided p-value for a t-statistic."""
    cdf = _t_cdf_approx(abs(t), df)
    return max(0.0, min(1.0, 2 * (1 - cdf)))


def _ols_simple(x: np.ndarray, y: np.ndarray) -> dict:
    """Pure-numpy simple linear regression. Returns slope, intercept, SE, p, R²."""
    n = len(x)
    if n < 3:
        return None
    x_mean = x.mean()
    y_mean = y.mean()
    ss_xx = ((x - x_mean) ** 2).sum()
    ss_xy = ((x - x_mean) * (y - y_mean)).sum()
    ss_yy = ((y - y_mean) ** 2).sum()

    if ss_xx == 0 or ss_yy == 0:
        return None

    slope = ss_xy / ss_xx
    intercept = y_mean - slope * x_mean
    r_sq = (ss_xy ** 2) / (ss_xx * ss_yy)

    # Residuals and standard error
    y_hat = intercept + slope * x
    residuals = y - y_hat
    sse = (residuals ** 2).sum()
    mse = sse / (n - 2) if n > 2 else 0
    se_slope = math.sqrt(mse / ss_xx) if ss_xx > 0 else 0

    # t-stat and p-value (two-sided)
    t_stat = slope / se_slope if se_slope > 0 else 0
    p_val = _two_sided_pvalue(t_stat, df=n - 2)

    return {
        "coefficient": float(slope),
        "intercept": float(intercept),
        "std_error": float(se_slope),
        "t_stat": float(t_stat),
        "p_value": float(p_val),
        "r_squared": float(r_sq),
        "n_obs": int(n),
    }


def run_iran_regression() -> dict:
    """Run OLS regression on Iran war data: daily volatility ~ conflict intensity.
    Pure-numpy implementation — no scipy required.
    """
    # Get data
    intensity = compute_iran_intensity()
    brent = _read_cache("brent_prices", 7200) or []

    if not intensity or len(brent) < 5:
        return {"error": "Insufficient data", "n_obs": 0}

    # Build Brent price map and compute daily volatility
    price_map = {b["date"]: b["price"] for b in brent}
    sorted_dates = sorted(price_map.keys())

    vol_map = {}
    for i in range(1, len(sorted_dates)):
        d = sorted_dates[i]
        prev = sorted_dates[i - 1]
        change = abs((price_map[d] - price_map[prev]) / price_map[prev]) * 100
        vol_map[d] = change

    intensity_map = {i["date"]: i["weekly_intensity"] for i in intensity}
    daily_intensity_map = {i["date"]: i["daily_intensity"] for i in intensity}

    war_dates = [d for d in sorted_dates if d >= "2026-02-28" and d in vol_map and d in intensity_map]

    if len(war_dates) < 5:
        return {"error": "Insufficient war-period observations", "n_obs": len(war_dates)}

    y = np.array([vol_map[d] for d in war_dates])
    x_weekly = np.array([intensity_map[d] for d in war_dates])
    x_daily = np.array([daily_intensity_map.get(d, 0) for d in war_dates])

    # Simple OLS models
    simple_weekly = _ols_simple(x_weekly, y)
    simple_daily = _ols_simple(x_daily, y)

    # Round for display
    def _round_result(r):
        if not r:
            return None
        return {k: (round(v, 6) if isinstance(v, float) else v) for k, v in r.items()}

    simple_result = _round_result(simple_weekly)
    daily_result = _round_result(simple_daily)

    # ── Multivariate OLS with DXY and OVX controls ──
    dxy_data = _read_cache("dxy", 7200) or []
    ovx_data = _read_cache("ovx", 7200) or []
    dxy_map = {d["date"]: d["value"] for d in dxy_data} if dxy_data else {}
    ovx_map = {d["date"]: d["value"] for d in ovx_data} if ovx_data else {}

    multi_dates = [d for d in war_dates if d in dxy_map and d in ovx_map]
    multi_result = None

    if len(multi_dates) >= 8:
        y_m = np.array([vol_map[d] for d in multi_dates])
        X = np.column_stack([
            np.ones(len(multi_dates)),
            [intensity_map[d] for d in multi_dates],
            [dxy_map[d] for d in multi_dates],
            [ovx_map[d] for d in multi_dates],
        ])

        try:
            betas, _, _, _ = np.linalg.lstsq(X, y_m, rcond=None)
            y_hat = X @ betas
            ss_res = np.sum((y_m - y_hat) ** 2)
            ss_tot = np.sum((y_m - np.mean(y_m)) ** 2)
            r_sq = 1 - ss_res / ss_tot if ss_tot > 0 else 0
            n, k = X.shape
            mse = ss_res / (n - k) if n > k else 0

            try:
                cov = mse * np.linalg.inv(X.T @ X)
                se = np.sqrt(np.diag(cov))
                t_stats = betas / se
                p_vals = [_two_sided_pvalue(float(t), df=n - k) for t in t_stats]
            except np.linalg.LinAlgError:
                se = [None] * k
                p_vals = [None] * k

            multi_result = {
                "labels": ["Intercept", "Iran Intensity", "DXY", "OVX"],
                "coefficients": [round(float(b), 6) for b in betas],
                "std_errors": [round(float(s), 6) if s is not None else None for s in se],
                "p_values": [round(float(p), 4) if p is not None else None for p in p_vals],
                "r_squared": round(float(r_sq), 4),
                "n_obs": int(n),
            }
        except Exception as e:
            logger.warning(f"Multivariate regression failed: {e}")

    return {
        "simple": simple_result,
        "daily": daily_result,
        "multivariate": multi_result,
        "war_dates": war_dates[:3] + ["..."] + war_dates[-3:] if len(war_dates) > 6 else war_dates,
        "date_range": f"{war_dates[0]} to {war_dates[-1]}",
        "caveat": f"Based on {len(war_dates)} trading days. Small sample — interpret with caution. Thesis used 505 observations over 24 months.",
    }


def get_hypothesis_results() -> dict:
    """Return hypothesis test results from the thesis defense regression table.
    Source: ThesisDefense_Hamm.pptx Slide 24 — OLS with HAC standard errors,
    7 control variables (DXY, OVX, SPR, OPEC Dummy, Russia-Ukraine Dummy,
    China PMI, Baker Hughes Rigs). 505 observations.
    """
    return {
        "h1": {
            "name": "H1: Attack Frequency",
            "description": "Higher frequency of Houthi maritime attacks increases Brent crude oil price volatility",
            "coefficient": -0.033,
            "p_value": 0.026,
            "r_squared": 0.487,
            "supported": False,
            "conclusion": "NOT SUPPORTED. The coefficient is statistically significant but negative (β = −0.033, p = 0.026), meaning more attacks are associated with less volatility — the opposite of the fear premium hypothesis. This supports market adaptation: as attacks became routine, markets priced in the disruption rather than reacting with increased volatility."
        },
        "h2": {
            "name": "H2: Tanker Specificity",
            "description": "Attacks specifically targeting oil tankers have a greater impact on volatility than general maritime attacks",
            "coefficient": -0.233,
            "p_value": 0.004,
            "r_squared": 0.464,
            "supported": False,
            "conclusion": "NOT SUPPORTED. Tanker-specific attacks show the strongest adaptation effect (β = −0.233, p = 0.004). Despite targeting energy infrastructure directly, tanker attacks are associated with reduced volatility — markets specifically recalibrated to energy-sector targeting."
        },
        "h3": {
            "name": "H3: Chokepoint Geography",
            "description": "Attacks at the Bab el-Mandeb strait chokepoint have a disproportionate impact on oil price volatility",
            "coefficient": -0.128,
            "p_value": 0.250,
            "r_squared": 0.459,
            "supported": False,
            "conclusion": "NOT SUPPORTED. The chokepoint coefficient is not statistically significant (β = −0.128, p = 0.250). Despite 67% of attacks concentrating in a 200km corridor near Bab el-Mandeb, geographic proximity alone does not predict oil price volatility."
        },
        "garch_summary": {
            "model": "GJR-GARCH(1,1,1)",
            "distribution": "Normal",
            "mean_model": "Constant",
            "observations": 505,
            "log_likelihood": -1027.56,
            "aic": 2065,
            "bic": 2086,
        },
        "model_comparison": {
            "labels": ["H1 (Attack Freq)", "H2 (Tanker)", "H3 (Chokepoint)"],
            "r_squared": [0.487, 0.464, 0.459],
            "finding": "With all 7 control variables and GARCH conditional variance as the dependent variable, models explain approximately 46–49% of volatility variance. H1 (p = 0.026) and H2 (p = 0.004) are statistically significant but negative — supporting market adaptation rather than a fear premium. H3 (p = 0.250) is not significant. Macroeconomic controls (OVX, DXY, Baker Hughes) dominate explanatory power."
        },
        "control_coefficients": {
            "h1": {
                "labels": ["Attack Freq (H1)", "DXY", "OVX", "SPR", "OPEC", "Russia-Ukraine", "China PMI", "Baker Hughes Rigs"],
                "coefficients": [-0.033, -0.134, 0.105, 0.000001, -0.318, -0.432, 0.173, 0.025],
                "std_errors": [0.015, 0.039, 0.013, 0.00000268, 0.256, 0.208, 0.162, 0.007],
                "p_values": [0.026, 0.001, 0.0001, 0.70, 0.21, 0.038, 0.29, 0.0004],
                "significant": [True, True, True, False, False, True, False, True]
            }
        }
    }
